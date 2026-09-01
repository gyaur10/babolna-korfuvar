import io
import re
import unicodedata

import pandas as pd
import streamlit as st
from openpyxl.styles import PatternFill

APP_VERSION = 'v4.8'
APP_RELEASE_DATE = '2026-09-01'
APP_CHANGELOG = {
    'v4.8 (2026-09-01)': 'A teljes 2025-12…2026-08 fuvarnapló (2 logbook, 7877 sor) '
                         'átnézése alapján, az ügyfél-kategóriák maradék eseteire: '
                         '(1) HARMADIK ORSZÁGOS ZÁRÁS: ha a kör időrendben utolsó lába '
                         '"Harmadik országba szállítás" típusú és nincs magyar állomása, '
                         'az nem hiányzó záró import, hanem szabályos, külföldön záruló '
                         'szolgáltatás → zöld, külön magyarázattal. '
                         '(2) IDŐSZAK ELŐTT KEZDŐDŐ KÖRÖK (⏮️ kék): a v4.6-os ablak-vég '
                         'logika tükörképe — export-nyitás nélküli kör az adatablak első '
                         'napjaiban, illetve Import törzs, amelynek csak a belföldi kiosztó '
                         'lába van meg. Nem hiba, a pénzügyi kimutatásokból kizárva; az '
                         'Összesítő külön bontja az ablak eleji és végi csonkolást. '
                         '(3) HARMADIK SORREND-VARIÁNS: a típus-validáció a részfeladat-, '
                         'a felvételi és a leadási idő szerinti láncot is végigpróbálja — '
                         'hiba csak akkor marad, ha egyik sorrend szerint sem konzisztens. '
                         'Dátum-anomáliás törzsnél (leadás < felvétel) a magyarázat ⚠️ '
                         'figyelmeztetést kap. '
                         '(4) SORREND-ELTÉRÉS LÁTHATÓVÁ TÉTELE: ha az időkapuk jók és a kör '
                         'kijön, a kör normálisan legenerálódik (zöld), csak figyelmeztetést '
                         'kap — ℹ️ mondat a Magyarázatban, új "Sorrend-figyelmeztetés" oszlop, '
                         'és a "Részfeladat-sorrend eltérés" munkalap Hatás oszlopa '
                         '(fals típushibát oldott fel / csak figyelmeztetés). '
                         '(5) TELJESÍTMÉNY: a kör-generálás ~2,5x gyorsabb (9 hónapos, '
                         '7877 soros adaton 23 mp → 9 mp). Törzsenkénti cache a '
                         'típus-validációra (eddig kétszer futott) és a láb-rekordokra, '
                         'lista-alapú rendezés a pandas sort_values helyett, memoizált '
                         'cím-feloldás, get_interval_with_addresses pandas nélkül. '
                         '(6) NAGY TÁBLA A FELÜLETEN: több hónap egyszerre több ezer sort '
                         'jelentett, a Styler HTML-je megfektette az oldalt (a letöltés '
                         'gombokig sem lehetett legörgetni). Mostantól hónaponként 150 sor '
                         'előnézet, checkboxszal kérhető a teljes tábla (színezés nélkül, '
                         'virtualizálva); a letöltött Excel mindig teljes.',
    'v4.7 (2026-09-01)': 'Ügyfél-ellenőrzés (Babolna_korfuvarfeladatok ellenőrzött) alapján: '
                         '(1) CÍM-NORMALIZÁLÁS újraírva — régi egybetűs, kötőjeles és szóköz '
                         'nélküli országkód (H-2800, D-31737, A 9020, DE54552), országnév a '
                         'cím bármely részén, és adatból tanult irányítószám+város → ország '
                         'szótár (AddressGazetteer). Megszűnt a hibás "4 jegyű irányítószám '
                         '= HU" szabály, amely a 4 jegyű irányítószámot használó országokat '
                         '(BE/LU/AT/CH/DK/NL) magyarnak vette. '
                         '(2) SORREND-ROBUSZTUS TÍPUS-VALIDÁCIÓ: ha a részfeladat-sorszámok '
                         'ellentmondanak az időkapuknak és az időrendi lánc konzisztens, '
                         'nincs többé fals típushiba. '
                         '(3) ÚJ OSZLOPOK: Fuvarfeladat típusok, Javasolt típus-javítás, '
                         'Kör kezdete hónap, Átnyúló kör. '
                         '(4) HÓNAPHATÁR: a besorolás marad Kör vége szerint, de az Összesítőn '
                         'és a Havi összesítőn látszik az előző hónapban indult körök száma és '
                         'az onnan áthozott bevétel. '
                         '(5) ÚJ MUNKALAPOK: Részfeladat-sorrend eltérés, Feloldatlan címek.',
    'v4.6 (2026-07-17)': 'Időszak után záródó körök felismerése: ha a kör / a hiányzó '
                         'nemzetközi részfeladat a feltöltött adatablak utolsó napjaira esik, '
                         'az nem rögzítési hiba, hanem a kör a következő időszakban fejeződik '
                         'be → kék info jelölés, és kizárás a pénzügyi kimutatásokból '
                         '(Összesítő, Ország-relációk, Megbízók, szumma sor).',
    'v4.5 (2026-07-15)': 'Feldolgozási animáció: guruló kamion + fázisonként frissülő '
                         'folyamatjelző a generálás alatt.',
    'v4.4 (2026-07-15)': 'Körfuvar-only mód: ha nincs költség / eredménykimutatás fájl '
                         'feltöltve, csak a körfuvar-generálás és a hibalisták futnak le '
                         '(aggregálás és profit-elemzés kihagyva) — gyorsabb hibajavítási kör.',
    'v4.3 (2026-07-15)': 'Több fuvarnapló feltöltés + dedup, év-hónap multiselect, '
                         'hónaponként külön Excel egy futásban, opcionális összesítő Excel, '
                         'eltérő tartalmú duplikátum figyelmeztetés.',
    'v4.2 (2026-07-15)': 'Komplex típus-/állomás-validáció közbenső állomásokkal (EU körfuvar '
                         'minta felismerés, javasolt javítással), szumma sor, Összesítő / '
                         'Ország-relációk / Megbízók / Típushibák / Dátum-anomáliák fülek.',
    'v4.1 (2026-07-15)': 'Kör-összefűzés törzs alapján (elő-fuvar kezelés), cím-normalizálás, '
                         'EU körfuvar irány, üres visszafutás jelzés, dátum-anomália warning.',
    'v4 (2026-07-15)':   'Útdíj dupla-számolás javítás, típus-validáció, részletes magyarázatok.',
}

st.set_page_config(page_title='Bábolna Körfuvar Generálás', layout='wide')
st.title('🚛 Bábolna Körfuvar Generálás')
st.caption(f'Verzió: **{APP_VERSION}** · Release: {APP_RELEASE_DATE}')
with st.expander('ℹ️ Verziótörténet'):
    for _ver, _desc in APP_CHANGELOG.items():
        st.markdown(f'- **{_ver}**: {_desc}')
st.markdown('---')

HU_PREFIX = 'HU '
BABOLNA_KEYWORD = 'Bábolna Rákóczi utca'

# Időszak-vég (adatablak-csonkolás) felismerés: ha egy félbeszakadtnak tűnő kör
# (nincs záró import) vagy egy hiányzó nemzetközi részfeladatú törzs utolsó
# aktivitása a feltöltött adatok legutolsó leadási dátumához képest ennyi napon
# belül van, akkor azt NEM hibaként, hanem "időszak után záródó kör (valószínű)"
# kategóriaként kezeljük (kék jelölés), és kizárjuk a pénzügyi kimutatásokból.
WINDOW_TRUNCATION_DAYS = 7

# A képernyőn megjelenített kör-tábla maximális sorszáma hónaponként (v4.8).
# A pandas Styler soronként generál HTML-t; több hónap egyszerre több ezer sort
# jelentene, ami a böngészőt megfekteti (a letöltés gombokig sem lehet legörgetni).
# A TELJES tábla mindig ott van a letölthető Excelben.
UI_PREVIEW_ROWS = 150

# Az első költség fájl ország-alapú kategóriái == a Flight Controlling fájl 'Útdíj költség'
# járatonkénti bontása. Ha az FC 'Útdíj költség' oszlop is jelen van, a szummában csak
# egyet szabad számolni – különben a Teljes költség duplikálja az útdíjat.
COUNTRY_TOLL_CATEGORIES = {
    'Belga', 'Francia', 'Holland', 'Magyar', 'Német',
    'Olasz', 'Osztrák', 'Spanifer', 'Szlovák', 'Szlovén',
}

# ---------------------------------------------------------------------------
# Cím-normalizálás (v4.7) – ország-felismerés
#
# A fuvarnaplóban a címek jelentős része nem szabványos: hiányzik az országkód
# ('6500 Baja...'), régi/egybetűs kód szerepel ('H-2800', 'D-31737', 'A 9020'),
# szóköz nélküli kód ('DE54552'), vagy csak a cím VÉGÉN van ország
# ('..., 38855 Wernigerode, Németország').
#
# A v4.6-ig érvényes szabály – "4 jegyű irányítószám prefix nélkül → HU" –
# hamis HU-t adott a 4 jegyű irányítószámot használó országokra is
# (BE/LU/AT/CH/DK/NL), ami hamis típushibákat okozott (pl. '4000 Liége' →
# HU→HU Import lánc → "hiányzó nemzetközi részfeladat" fals riasztás).
#
# Az új feloldás sorrendje:
#   1. explicit országkód prefix (ISO2, régi egybetűs, kötőjeles, szóköz nélküli)
#   2. országnév a címben bárhol (magyar / angol / német / helyi alak)
#   3. adatból épített irányítószám+város → ország szótár (gazetteer):
#      a PREFIXES címekből tanulunk, és azt alkalmazzuk a prefix nélküliekre
#   4. utcanév-nyelv + irányítószám-hossz heurisztika (5 jegyű + 'Straße' → DE, …)
#   5. 4 jegyű irányítószám / magyar utcanév-jelölő, külföldi jel nélkül → HU
#   6. egyébként '??' (és bekerül a "Feloldatlan címek" riportba)
# ---------------------------------------------------------------------------
_CC_ISO2 = {
    'HU', 'AT', 'DE', 'FR', 'IT', 'NL', 'BE', 'ES', 'SK', 'SI', 'PL', 'CZ', 'RO', 'HR',
    'CH', 'GB', 'UK', 'SE', 'DK', 'PT', 'LU', 'LT', 'LV', 'EE', 'BG', 'GR', 'FI', 'IE',
    'NO', 'RS', 'BA', 'MK', 'TR', 'UA', 'MD', 'AL', 'ME', 'LI', 'MT', 'CY', 'IS',
}
_CC_ALIAS = {
    'UK': 'GB', 'H': 'HU', 'D': 'DE', 'A': 'AT', 'I': 'IT', 'F': 'FR', 'B': 'BE',
    'E': 'ES', 'P': 'PT', 'L': 'LU', 'N': 'NO', 'S': 'SE', 'M': 'MT', 'FIN': 'FI',
    'SLO': 'SI', 'IRL': 'IE', 'SRB': 'RS', 'BIH': 'BA', 'EST': 'EE',
}

# Visszafelé kompatibilitás: a régi _CC_RE-t más kód is használhatja
_CC_RE = re.compile(
    r'^(HU|AT|DE|FR|IT|NL|BE|ES|SK|SI|PL|CZ|RO|HR|CH|GB|UK|SE|DK|PT|LU|LT|LV|EE|BG|GR|FI|IE|NO)\s',
    re.IGNORECASE,
)

_COUNTRY_NAMES = {
    'magyarorszag': 'HU', 'hungary': 'HU', 'ungarn': 'HU', 'hongrie': 'HU',
    'nemetorszag': 'DE', 'germany': 'DE', 'deutschland': 'DE', 'allemagne': 'DE',
    'ausztria': 'AT', 'austria': 'AT', 'osterreich': 'AT', 'autriche': 'AT',
    'olaszorszag': 'IT', 'italy': 'IT', 'italia': 'IT', 'italien': 'IT',
    'franciaorszag': 'FR', 'france': 'FR', 'frankreich': 'FR',
    'belgium': 'BE', 'belgique': 'BE', 'belgie': 'BE', 'belgien': 'BE',
    'hollandia': 'NL', 'netherlands': 'NL', 'nederland': 'NL', 'niederlande': 'NL',
    'csehorszag': 'CZ', 'czechia': 'CZ', 'tschechien': 'CZ',
    'szlovakia': 'SK', 'slovakia': 'SK', 'slovensko': 'SK', 'slowakei': 'SK',
    'lengyelorszag': 'PL', 'poland': 'PL', 'polska': 'PL', 'polen': 'PL',
    'romania': 'RO', 'rumanien': 'RO',
    'spanyolorszag': 'ES', 'spain': 'ES', 'espana': 'ES', 'spanien': 'ES',
    'svajc': 'CH', 'switzerland': 'CH', 'schweiz': 'CH', 'suisse': 'CH',
    'szlovenia': 'SI', 'slovenia': 'SI', 'slowenien': 'SI',
    'horvatorszag': 'HR', 'croatia': 'HR', 'hrvatska': 'HR', 'kroatien': 'HR',
    'luxemburg': 'LU', 'luxembourg': 'LU',
    'portugalia': 'PT', 'portugal': 'PT',
    'dania': 'DK', 'denmark': 'DK', 'danmark': 'DK',
    'svedorszag': 'SE', 'sweden': 'SE', 'sverige': 'SE',
    'egyesult kiralysag': 'GB', 'united kingdom': 'GB', 'anglia': 'GB', 'england': 'GB',
    'irorszag': 'IE', 'ireland': 'IE',
    'szerbia': 'RS', 'serbia': 'RS', 'torokorszag': 'TR', 'turkey': 'TR',
    'ukrajna': 'UA', 'ukraine': 'UA', 'finnorszag': 'FI', 'finland': 'FI',
    'norvegia': 'NO', 'norway': 'NO', 'gorogorszag': 'GR', 'greece': 'GR',
    'bulgaria': 'BG', 'lithuania': 'LT', 'latvia': 'LV', 'estonia': 'EE',
}

# Városnév-kinyeréskor kihagyandó általános tokenek (utcatípus, cégforma, épület)
_ADDR_STOPWORDS = {
    'utca', 'ter', 'krt', 'korut', 'koz', 'dulo', 'hrsz', 'ipari', 'zona', 'ipartelep',
    'street', 'road', 'str', 'strasse', 'strase', 'weg', 'allee', 'platz', 'gasse',
    'ring', 'damm', 'via', 'viale', 'strada', 'corso', 'piazza', 'localita', 'loc',
    'rue', 'avenue', 'boulevard', 'chemin', 'route', 'impasse', 'laan', 'straat',
    'dreef', 'calle', 'poligono', 'carretera', 'ctra',
    'gmbh', 'kft', 'zrt', 'sarl', 'spa', 'srl', 'plc', 'ltd', 'kgaa',
    'des', 'les', 'del', 'della', 'van', 'der', 'den', 'und', 'and',
    'sur', 'sous', 'saint', 'ste', 'auf', 'bei', 'zum', 'zur',
    'nord', 'sud', 'est', 'ouest', 'ovest', 'ost', 'west', 'sued',
    'industrial', 'industrie', 'industriegebiet', 'gewerbegebiet',
    'logistics', 'logistik', 'terminal',
}
_ADDR_LETTERS_RE = re.compile(r'[^a-z ]+')
_ADDR_PC_RE = re.compile(r'(?<!\d)(\d{4,6})(?!\d)')

# Utcanév-nyelv jelölők (irányítószám-hosszal kombinálva ad országot)
_LANG_DE_RE = re.compile(r'(strasse|straße|\bstr\.|\bstr\b|\bweg\b|\bgasse\b|platz|allee|anschlussstelle|gewerbegebiet|\bwerk\b)', re.I)
_LANG_FR_RE = re.compile(r'(\brue\b|avenue|boulevard|\bchemin\b|\broute\b|impasse|zone industrielle|\bz\.i\.)', re.I)
_LANG_IT_RE = re.compile(r'(\bvia\b|viale|strada|corso|piazza|localita|località)', re.I)
_LANG_NL_RE = re.compile(r'(straat|\blaan\b|dreef)', re.I)
_LANG_ES_RE = re.compile(r'(\bcalle\b|carretera|poligono|polígono)', re.I)

_HU_MARK_RE = re.compile(r'(\butca\b|\bu\.|\bút\b|\but\.|\btér\b|\bkrt\b|körút|\bköz\b|dűlő|hrsz|ipari park|puszta|major)', re.I)
_FOREIGN_MARK_RE = re.compile(
    r'(\brue\b|avenue|boulevard|chemin|\broute\b|strasse|straße|\bstr\.|\bweg\b|\bgasse\b'
    r'|platz|allee|\bvia\b|viale|strada|corso|piazza|calle|carretera|\bstreet\b|\broad\b'
    r'|\blaan\b|straat|dreef)', re.I)


def _addr_fold(s) -> str:
    """Ékezet nélküli kisbetűs alak (unicode NFKD)."""
    s = unicodedata.normalize('NFKD', str(s).lower())
    s = ''.join(ch for ch in s if not unicodedata.combining(ch))
    return s.replace('ß', 'ss')


def _addr_prefix_cc(c: str):
    """Explicit országkód a cím elejéről ('HU 2943', 'D-31737', 'A 9020', 'DE54552')."""
    m = (re.match(r'^([A-Za-z]{1,3})\s*[-–]\s*\d', c)
         or re.match(r'^([A-Za-z]{1,3})\s+\d', c)
         or re.match(r'^([A-Za-z]{2,3})(\d{4,6})\b', c)
         or re.match(r'^([A-Za-z]{2,3})\s+[A-Za-zÀ-ž]', c))
    if not m:
        return None
    code = m.group(1).upper()
    code = _CC_ALIAS.get(code, code)
    return code if code in _CC_ISO2 else None


# Egyetlen előfordított alternáció a ~60 országnévre (60 külön re.search helyett)
_COUNTRY_NAME_RE = re.compile(
    r'\b(' + '|'.join(sorted((re.escape(n) for n in _COUNTRY_NAMES), key=len, reverse=True)) + r')\b'
)


def _addr_country_name_cc(c: str):
    """Országnév a címben — az UTOLSÓ előfordulás számít.

    A cím végén álló ország a mérvadó, nem a cégnévben szereplő: pl.
    'JRS-Austria-raktár, Papírgyári út 42., 2400 DUNAÚJVÁROS, HUNGARY' → HU, nem AT.
    """
    last = None
    for m in _COUNTRY_NAME_RE.finditer(_addr_fold(c)):
        last = m
    return _COUNTRY_NAMES[last.group(1)] if last else None


def _addr_tokens(c: str):
    folded = _addr_fold(c)
    return [t for t in _ADDR_LETTERS_RE.sub(' ', folded).split()
            if len(t) >= 3 and t not in _ADDR_STOPWORDS]


class AddressGazetteer:
    """Irányítószám+város → ország szótár, a saját adatból tanulva.

    Az explicit országkóddal rögzített címekből (a címek ~95%-a) megtanuljuk,
    hogy egy (irányítószám, városnév) vagy csak városnév melyik országhoz
    tartozik, és ezt alkalmazzuk a prefix nélküli címekre.
    """

    def __init__(self):
        self.pc_city = {}
        self.city = {}
        self.pc = {}

    @staticmethod
    def _bump(store, key, cc):
        d = store.setdefault(key, {})
        d[cc] = d.get(cc, 0) + 1

    def add(self, cim, cc):
        toks = _addr_tokens(cim)
        pcs = _ADDR_PC_RE.findall(str(cim))
        for t in toks:
            self._bump(self.city, t, cc)
        for p in pcs:
            self._bump(self.pc, p, cc)
            for t in toks:
                self._bump(self.pc_city, (p, t), cc)

    def build(self, cimek):
        """Két menet: előbb az explicit prefixes címek, majd az országnévvel
        azonosítottak (így a '…, Ausztria' variáns megtanítja a prefix nélkülit)."""
        cimek = [str(c).strip() for c in cimek if str(c).strip()]
        rest = []
        for c in cimek:
            cc = _addr_prefix_cc(c)
            if cc:
                self.add(c, cc)
            else:
                rest.append(c)
        for c in rest:
            cc = _addr_country_name_cc(c)
            if cc:
                self.add(c, cc)
        return self

    @staticmethod
    def _best(counts):
        if not counts:
            return None
        items = sorted(counts.items(), key=lambda kv: -kv[1])
        if len(items) == 1 or items[0][1] >= 3 * items[1][1]:
            return items[0][0]
        return None

    def lookup(self, cim):
        c = str(cim).strip()
        toks = _addr_tokens(c)
        pcs = _ADDR_PC_RE.findall(c)
        for p in pcs:
            for t in toks:
                r = self._best(self.pc_city.get((p, t)))
                if r:
                    return r
        votes = {}
        for t in toks:
            r = self._best(self.city.get(t))
            if r:
                votes[r] = votes.get(r, 0) + 1
        if votes:
            top = sorted(votes.items(), key=lambda kv: -kv[1])
            if len(top) == 1 or top[0][1] > top[1][1]:
                return top[0][0]
        for p in pcs:
            r = self._best(self.pc.get(p))
            if r:
                return r
        return None


# Modul-szintű gazetteer: a pipeline elején töltjük fel a teljes fuvarnaplóból.
_ADDR_GAZETTEER = AddressGazetteer()


def build_address_gazetteer(df: pd.DataFrame,
                            columns=('Első Felvételi állomás cím',
                                     'Utolsó Leadási állomás cím')):
    """A modul-szintű gazetteer (újra)építése a fuvarnapló címeiből."""
    global _ADDR_GAZETTEER
    vals = []
    for c in columns:
        if c in df.columns:
            vals.extend(df[c].dropna().astype(str).unique().tolist())
    _ADDR_GAZETTEER = AddressGazetteer().build(vals)
    _COUNTRY_CACHE.clear()
    return _ADDR_GAZETTEER


def _addr_lang_cc(c: str):
    """Utcanév-nyelv + irányítószám-hossz heurisztika (csak prefix/gazetteer után)."""
    pcs = _ADDR_PC_RE.findall(c)
    plen = len(pcs[0]) if pcs else 0
    if plen == 5:
        if _LANG_DE_RE.search(c):
            return 'DE'
        if _LANG_IT_RE.search(c):
            return 'IT'
        if _LANG_FR_RE.search(c):
            return 'FR'
        if _LANG_ES_RE.search(c):
            return 'ES'
    if plen == 4:
        if _LANG_NL_RE.search(c):
            return 'NL'
        if _LANG_DE_RE.search(c):
            return 'AT'
    return None


_COUNTRY_CACHE = {}


def country_of(cim) -> str:
    """Egy állomás-cím országkódja ('HU', 'DE', … vagy '??' ha nem felismerhető).

    Teljesítmény (v4.8): a feloldás eredménye címenként cache-elve — a fuvarnaplóban
    ~7000 cím-előfordulásra ~1000 egyedi cím jut, és a függvény a legforgalmasabb
    hívási útvonalon van (is_hu_address → irány-osztályozás, típus-validáció).
    A cache a gazetteer újraépítésekor ürül.
    """
    c = str(cim or '').strip()
    if not c:
        return '??'
    hit = _COUNTRY_CACHE.get(c)
    if hit is not None:
        return hit
    _COUNTRY_CACHE[c] = res = _country_of_uncached(c)
    return res


def _country_of_uncached(c: str) -> str:
    cc = _addr_prefix_cc(c)
    if cc:
        return cc
    cc = _addr_country_name_cc(c)
    if cc:
        return cc
    cc = _ADDR_GAZETTEER.lookup(c)
    if cc:
        return cc
    cc = _addr_lang_cc(c)
    if cc:
        return cc
    pcs = _ADDR_PC_RE.findall(c)
    if pcs and len(pcs[0]) == 4 and not _FOREIGN_MARK_RE.search(c):
        return 'HU'
    if _HU_MARK_RE.search(c) and not _FOREIGN_MARK_RE.search(c):
        return 'HU'
    return '??'


def is_hu_address(cim) -> bool:
    """Egy állomás-cím Magyarországra mutat-e (a country_of feloldás alapján)."""
    return country_of(cim) == 'HU'


def unresolved_addresses(df: pd.DataFrame,
                         columns=('Első Felvételi állomás cím',
                                  'Utolsó Leadási állomás cím')) -> pd.DataFrame:
    """A '??' országú (fel nem oldható) címek riportja – forrásrendszeri javításhoz."""
    rows = {}
    for col in columns:
        if col not in df.columns:
            continue
        for cim in df[col].dropna().astype(str):
            c = cim.strip()
            if c and country_of(c) == '??':
                rows[c] = rows.get(c, 0) + 1
    if not rows:
        return pd.DataFrame(columns=['Cím', 'Előfordulás'])
    out = pd.DataFrame(sorted(rows.items(), key=lambda kv: -kv[1]),
                       columns=['Cím', 'Előfordulás'])
    return out



# ---------------------------------------------------------------------------
# Segéd függvények – Fuvarszám bontás
# ---------------------------------------------------------------------------
def _torzs_of(f_szam: str) -> str:
    """Fuvarszám-törzs a Fuvarszám-ból (az első '-' előtti rész)."""
    return str(f_szam).split('-')[0]


def _reszfeladat_of(f_szam):
    """Részfeladat-sorszám (int) a Fuvarszám-ból, vagy None, ha nem számszerű / nincs."""
    parts = str(f_szam).split('-')
    if len(parts) < 2:
        return None
    try:
        return int(parts[1])
    except (ValueError, TypeError):
        return None


# ---------------------------------------------------------------------------
# Egy fuvarrészfeladat irányának osztályozása
# ---------------------------------------------------------------------------
def classify_leg_direction(row):
    """Egy részfeladat irányának meghatározása.

    Visszatér: kifelé-nemzetközi / kifelé-belföldi / befelé-nemzetközi /
               befelé-belföldi / semleges / korfuvar / ismeretlen.

    v4.1: az 'EU körfuvar' típus külön ágat kapott. Egy HU→HU EU körfuvar láb
    oda-vissza utat ír le egyetlen sorban (tipikusan Bábolna→Bábolna), ezért
    'korfuvar' irányt kap, ami a kör tartalmában kifelé ÉS befelé szakasznak
    is számít (önmagában teljes kört ad).
    """
    fel = str(row['Első Felvételi állomás cím'])
    le = str(row['Utolsó Leadási állomás cím'])
    tipus = str(row['Fuvarfeladat típusa'])

    fel_hu = is_hu_address(fel)
    le_hu = is_hu_address(le)
    fel_babolna = BABOLNA_KEYWORD in fel
    le_babolna = BABOLNA_KEYWORD in le

    # EU körfuvar – a típus-alapú Export/Import ágak ELŐTT kell vizsgálni
    tl = tipus.lower()
    if 'eu' in tl and 'körfuvar' in tl:
        if fel_hu and le_hu:
            return 'korfuvar'
        if fel_hu and not le_hu:
            return 'kifelé-nemzetközi'
        if not fel_hu and le_hu:
            return 'befelé-nemzetközi'
        return 'semleges'

    # Típus-alapú „gyors" osztályozás
    if 'Export' in tipus:
        return 'kifelé-nemzetközi'
    if 'Import' in tipus:
        return 'befelé-nemzetközi'
    if tipus.startswith('Harmadik országba szállítás'):
        return 'semleges'

    # Címek alapján történő finomhangolás
    if fel_hu and not fel_babolna and le_babolna:
        return 'kifelé-belföldi'
    if fel_babolna and le_hu and not le_babolna:
        return 'kifelé-belföldi'
    if fel_babolna and not le_hu:
        return 'kifelé-nemzetközi'
    if not fel_hu and le_babolna:
        return 'befelé-nemzetközi'
    if fel_babolna and le_hu and not le_babolna:
        return 'befelé-belföldi'

    return 'ismeretlen'


# ---------------------------------------------------------------------------
# Egy fuvarszám-törzs kezdő / záró sora
# ---------------------------------------------------------------------------
def _sort_legs_chronologically(group: pd.DataFrame) -> pd.DataFrame:
    """Egy törzs (vagy kör) lábainak sorrendje IDŐREND szerint (v4.7).

    v4.6-ig a rendezés a Fuvarszám részfeladat-sorszáma szerint történt. Az ügyfél
    ellenőrzése kimutatta, hogy a részfeladat-sorszámok jelentős részben fel vannak
    cserélve (tipikusan a 2. és 3. részfeladat), az IDŐKAPUK viszont helyesek.
    A sorszám szerinti rendezés így hibás kezdő/záró lábat választott, ami hamis
    "hibás fuvarfeladat típus" riasztásokat okozott.

    Elsődleges kulcs: első felvételi időkapu, majd utolsó leadási időkapu,
    végül – holtverseny esetén – a részfeladat-sorszám.
    """
    g = group.copy()
    g['_reszfeladat'] = g['Fuvarszám'].map(_reszfeladat_of)
    keys = [c for c in ('Első Felvételi állomás időkapu (dátum)',
                        'Utolsó Leadási állomás időkapu (dátum)') if c in g.columns]
    if keys and g[keys[0]].notna().any():
        return g.sort_values(keys + ['_reszfeladat'], na_position='last')
    return g.sort_values('_reszfeladat', na_position='last')


def _sort_legs_by_leadas(group: pd.DataFrame) -> pd.DataFrame:
    """Lábak sorrendje az UTOLSÓ LEADÁSI időkapu szerint (v4.7).

    A kör-építés maga is leadási dátum szerint rendez; dátum-anomáliás törzsnél
    (leadás < felvétel) ez a sorrend gyakran helyesebb, mint a felvételi szerinti.
    """
    g = group.copy()
    if '_reszfeladat' not in g.columns:
        g['_reszfeladat'] = g['Fuvarszám'].map(_reszfeladat_of)
    keys = [c for c in ('Utolsó Leadási állomás időkapu (dátum)',
                        'Első Felvételi állomás időkapu (dátum)') if c in g.columns]
    if keys and g[keys[0]].notna().any():
        return g.sort_values(keys + ['_reszfeladat'], na_position='last')
    return g.sort_values('_reszfeladat', na_position='last')


def _van_datum_anomalia(group: pd.DataFrame) -> bool:
    """Van-e a törzsben olyan láb, ahol a leadási időkapu korábbi a felvételinél."""
    fel = pd.to_datetime(group.get('Első Felvételi állomás időkapu (dátum)'), errors='coerce')
    le = pd.to_datetime(group.get('Utolsó Leadási állomás időkapu (dátum)'), errors='coerce')
    if fel is None or le is None:
        return False
    return bool((le.notna() & fel.notna() & (le < fel)).any())


def _reszfeladat_sorrend_elter(group: pd.DataFrame, torzs=None) -> bool:
    """Igaz, ha a részfeladat-sorszámok sorrendje eltér az időrendtől."""
    if torzs is None:
        torzs = _torzs_of(str(group['Fuvarszám'].iloc[0])) if len(group) else ''
    recs = _order_records(_leg_records(torzs, group), 'chrono')
    nums = [l['_rf_raw'] for l in recs if l['_rf_raw'] is not None]
    return len(nums) > 1 and nums != sorted(nums)


def build_reszfeladat_sorrend_table(df: pd.DataFrame) -> pd.DataFrame:
    """Riport azokról a törzsekről, ahol a részfeladat-sorszám ellentmond az időrendnek.

    Nem hiba a kör szempontjából (a generálás időrend szerint dolgozik), de a
    forrásrendszerben javítandó adatminőségi jelzés.
    """
    rows = []
    tmp = df.assign(_torzs=df['Fuvarszám'].astype(str).map(_torzs_of))
    for torzs, grp in tmp.groupby('_torzs'):
        if len(grp) < 2:
            continue
        ordered = _sort_legs_chronologically(grp)
        nums = [n for n in ordered['_reszfeladat'].tolist() if pd.notna(n)]
        if len(nums) > 1 and nums != sorted(nums):
            rows.append({
                'Törzs': torzs,
                'Hatás': ('feloldotta a fals típushibát — a kör hibátlanul legenerálva'
                          if str(torzs) in _ORDER_RESOLVED_TORZSEK
                          else 'csak figyelmeztetés — a kör enélkül is hibátlan volt'),
                'Részfeladatok időrendben': ', '.join(
                    str(f) for f in ordered['Fuvarszám'].astype(str)),
                'Sorszám-sorrend': ' → '.join(str(int(n)) for n in nums),
                'Járatszámok': ', '.join(dict.fromkeys(
                    ordered['Járatszám'].astype(str))),
                'Első felvétel': ordered.iloc[0].get('Első Felvételi állomás időkapu (dátum)'),
                'Utolsó leadás': ordered.iloc[-1].get('Utolsó Leadási állomás időkapu (dátum)'),
                'Javasolt javítás': 'A részfeladat-sorszámok felcserélve — '
                                    'a forrásrendszerben az időkapuk szerinti '
                                    'sorrendre javítandók.',
            })
    if not rows:
        return pd.DataFrame(columns=['Törzs', 'Hatás', 'Részfeladatok időrendben',
                                     'Sorszám-sorrend', 'Járatszámok', 'Első felvétel',
                                     'Utolsó leadás', 'Javasolt javítás'])
    return pd.DataFrame(rows).sort_values('Törzs').reset_index(drop=True)


def _torzs_start_end(group: pd.DataFrame):
    """Egy adott fuvarszám-törzs csoportjára visszaadja a törzs kezdő és záró sorát.
    - Ha van értelmezhető részfeladat szám (-1, -2, ...), akkor a legkisebb részfeladatszámú
      sor a kezdő, a legnagyobb részfeladatszámú a záró.
    - Ha nincs részfeladat szám, időrendi min/max lesz belőle.
    Visszatér: (row_start, row_end) vagy (None, None).
    """
    if group is None or group.empty:
        return None, None

    group = group.copy()
    group['_reszfeladat'] = group['Fuvarszám'].map(_reszfeladat_of)

    has_reszf = group['_reszfeladat'].notna()
    if has_reszf.any():
        sub = group[has_reszf]
        if len(sub) > 1:
            row_start = sub.loc[sub['_reszfeladat'].idxmin()]
            row_end = sub.loc[sub['_reszfeladat'].idxmax()]
            return row_start, row_end
        only = sub.iloc[0]
        return only, only

    valid_start = group.dropna(subset=['Első Felvételi állomás időkapu (dátum)'])
    valid_end = group.dropna(subset=['Utolsó Leadási állomás időkapu (dátum)'])
    if valid_start.empty or valid_end.empty:
        return None, None

    row_start = valid_start.loc[valid_start['Első Felvételi állomás időkapu (dátum)'].idxmin()]
    row_end = valid_end.loc[valid_end['Utolsó Leadási állomás időkapu (dátum)'].idxmax()]
    return row_start, row_end


def get_interval_with_addresses(legs_df: pd.DataFrame):
    """Adott részfeladat-halmazhoz (pl. egy kör "kifelé" sorai) visszaadja a
    (kezdő időkapu, kezdő cím, záró időkapu, záró cím) tuple-t.

    Minden benne szereplő fuvarszám-törzset KÜLÖN kezel: törzsenként meghatározza a
    törzs saját kezdetét és végét (_torzs_start_end), majd az összes törzs kezdő pontjai
    közül a legkorábbit, a záró pontok közül a legkésőbbit választja.
    """
    if legs_df is None or legs_df.empty:
        return pd.NaT, None, pd.NaT, None

    needed_cols = [
        'Fuvarszám',
        'Első Felvételi állomás időkapu (dátum)',
        'Utolsó Leadási állomás időkapu (dátum)',
        'Első Felvételi állomás cím',
        'Utolsó Leadási állomás cím',
    ]
    for c in needed_cols:
        if c not in legs_df.columns:
            return pd.NaT, None, pd.NaT, None

    # Teljesítmény (v4.8): oszlop-listákon dolgozunk, pandas copy/groupby nélkül —
    # ez a függvény körönként 3-4x hívódik (kifelé/semleges/befelé szakaszokra).
    fsz = legs_df['Fuvarszám'].astype(str).tolist()
    fdt = legs_df['Első Felvételi állomás időkapu (dátum)'].tolist()
    ldt = legs_df['Utolsó Leadási állomás időkapu (dátum)'].tolist()
    fad = legs_df['Első Felvételi állomás cím'].tolist()
    lad = legs_df['Utolsó Leadási állomás cím'].tolist()

    per_torzs = {}
    for i, f in enumerate(fsz):
        per_torzs.setdefault(_torzs_of(f), []).append(i)

    start_candidates = []
    end_candidates = []
    # A törzsek bejárása rendezetten (mint a korábbi pandas groupby), hogy az
    # időkapu-holtversenyeknél ugyanaz a cím nyerjen, mint eddig.
    for _t in sorted(per_torzs):
        idxs = per_torzs[_t]
        rf = [(i, _reszfeladat_of(fsz[i])) for i in idxs]
        with_rf = [(i, n) for i, n in rf if n is not None]
        if with_rf:
            if len(with_rf) > 1:
                i_start = min(with_rf, key=lambda x: x[1])[0]
                i_end = max(with_rf, key=lambda x: x[1])[0]
            else:
                i_start = i_end = with_rf[0][0]
        else:
            cand_s = [i for i in idxs if pd.notna(fdt[i])]
            cand_e = [i for i in idxs if pd.notna(ldt[i])]
            if not cand_s or not cand_e:
                continue
            i_start = min(cand_s, key=lambda i: fdt[i])
            i_end = max(cand_e, key=lambda i: ldt[i])
        if pd.notna(fdt[i_start]):
            start_candidates.append((fdt[i_start], fad[i_start]))
        if pd.notna(ldt[i_end]):
            end_candidates.append((ldt[i_end], lad[i_end]))

    if not start_candidates or not end_candidates:
        return pd.NaT, None, pd.NaT, None

    start_dt, start_addr = min(start_candidates, key=lambda x: x[0])
    end_dt, end_addr = max(end_candidates, key=lambda x: x[0])
    return start_dt, start_addr, end_dt, end_addr


# ---------------------------------------------------------------------------
# Fuvarfeladat típus – elvárt viszonylat és validáció
# ---------------------------------------------------------------------------
def _classify_fuvarfeladat_type_expected(tipus_str) -> str:
    """A 'Fuvarfeladat típusa' oszlopból meghatározza az elvárt viszonylatot.
    Visszatér: 'EU_KOR' / 'EXPORT' / 'IMPORT' / 'SEMLEGES' / 'ISMERETLEN'.
    Sorrend fontos: 'EU körfuvar' előbb, mert egyben export/import szó is szerepelhet."""
    t = str(tipus_str or '')
    tl = t.lower()
    if ('eu' in tl) and ('körfuvar' in tl or 'koerfuvar' in tl or 'kör fuvar' in tl):
        return 'EU_KOR'
    if 'export' in tl:
        return 'EXPORT'
    if 'import' in tl:
        return 'IMPORT'
    if 'harmadik ország' in tl or 'harmadik orszag' in tl:
        return 'SEMLEGES'
    return 'ISMERETLEN'


# Teljesítmény (v4.8): a törzsenkénti láb-rekordok EGYSZER készülnek el (pandas
# iterrows + cím-feloldás), a három lehetséges sorrend utána már csak listát rendez.
# Korábban minden sorrend-variáns külön DataFrame copy+sort_values volt, ami a
# futásidő felét vitte (16 000+ sort_values hívás egy 2200 körös futáson).
_TORZS_LEGS_CACHE = {}
_TORZS_ANALYSIS_CACHE = {}

_DT_MIN = pd.Timestamp.min
_DT_MAX = pd.Timestamp.max


def _leg_records(torzs, torzs_group: pd.DataFrame):
    """Egy törzs lábai plain dict-ekként, rendezési kulcsokkal. Cache-elt."""
    key = str(torzs)
    hit = _TORZS_LEGS_CACHE.get(key)
    if hit is not None:
        return hit
    n = len(torzs_group)
    cols = torzs_group.columns

    def col(name, default=None):
        if name in cols:
            return torzs_group[name].tolist()
        return [default] * n

    c_fsz = [str(x) for x in col('Fuvarszám', '')]
    c_jsz = [str(x) for x in col('Járatszám', '')]
    c_fel = col('Első Felvételi állomás cím', '')
    c_le = col('Utolsó Leadási állomás cím', '')
    c_dij = col('Díj részarány (EUR)', 0)
    c_tip = col('Fuvarfeladat típusa', '')
    c_fdt = col('Első Felvételi állomás időkapu (dátum)')
    c_ldt = col('Utolsó Leadási állomás időkapu (dátum)')
    c_von = col('Vontatmány', '')

    recs = []
    for i in range(n):
        fel = str(c_fel[i] or '').strip()
        le = str(c_le[i] or '').strip()
        rf = _reszfeladat_of(c_fsz[i])
        fel_dt, le_dt = c_fdt[i], c_ldt[i]
        try:
            dij = float(c_dij[i])
            if dij != dij:  # NaN
                dij = 0.0
        except (TypeError, ValueError):
            dij = 0.0
        recs.append({
            'fsz': c_fsz[i],
            'jsz': c_jsz[i],
            'fel': fel, 'le': le,
            'fel_hu': country_of(fel) == 'HU', 'le_hu': country_of(le) == 'HU',
            'dij': dij,
            'tipus': str(c_tip[i] or ''),
            'vontatmany': str(c_von[i] or ''),
            '_rf': (1, 0) if rf is None else (0, rf),
            '_fel': (1, _DT_MIN) if pd.isna(fel_dt) else (0, pd.Timestamp(fel_dt)),
            '_le': (1, _DT_MIN) if pd.isna(le_dt) else (0, pd.Timestamp(le_dt)),
            '_rf_raw': rf,
            '_fel_raw': None if pd.isna(fel_dt) else pd.Timestamp(fel_dt),
            '_le_raw': None if pd.isna(le_dt) else pd.Timestamp(le_dt),
        })
    _TORZS_LEGS_CACHE[key] = recs
    return recs


def _order_records(recs, order):
    """Láb-rekordok rendezése a három lehetséges kulcs szerint (lista-rendezés)."""
    if order == 'chrono':
        return sorted(recs, key=lambda r: (r['_fel'], r['_le'], r['_rf']))
    if order == 'leadas':
        return sorted(recs, key=lambda r: (r['_le'], r['_fel'], r['_rf']))
    return sorted(recs, key=lambda r: (r['_rf'], r['_fel']))


def _reset_torzs_caches():
    _TORZS_LEGS_CACHE.clear()
    _TORZS_ANALYSIS_CACHE.clear()
    _ORDER_RESOLVED_TORZSEK.clear()


def _analyze_torzs_type_ordered(torzs: str, torzs_group: pd.DataFrame,
                                window_end=None, order='reszfeladat',
                                window_start=None):
    """Komplex fuvarfeladat típus + állomás-típus validáció egy törzsre.

    A törzs ÖSSZES részfeladatát vizsgálja (részfeladat-sorszám szerinti sorrendben),
    a közbenső fel-/lerakó állomásokkal együtt. Így megkülönböztethető:
      - a valódi típushiba (pl. Export ami valójában Import),
      - az EU körfuvar mintázat Export/Import típussal ("Wolfsburg minta":
        van hurok-részfeladat, ahol felvétel == leadás, mert a hazafelé felrakó
        LERAKÓKÉNT lett rögzítve, ezért a külföldi állomás eltűnik a naplóból),
      - a hiányzó hazatérő szakasz (EU körfuvarnál külföldön áll meg a lánc),
      - az adatablak-csonkolás (a nemzetközi láb a lekérdezett időszakon kívül esik).

    Alapelv (raksúly hiányában): ha komplett kör van, a hazafelé tartó viszonylaton
    mindenképpen lennie kell egy FELRAKÓNAK, amit itthon le lehet rakni. Ha nincs,
    feltételezhetően állomás-típus hiba van (lerakó felrakás helyett).

    Visszatér: None (nincs hiba) vagy dict:
      {'torzs','tipus','viszonylat','fuvarszamok','jaratszamok',
       'hiba','javaslat','implies_complete'}
    implies_complete=True → a kör a valóságban lezárult (ki- és hazaút megvolt),
    csak a rögzítés hibás; a kör tartalmilag teljesnek tekinthető.
    """
    if torzs_group is None or torzs_group.empty:
        return None

    legs = _order_records(_leg_records(torzs, torzs_group), order)
    if not legs:
        return None

    first, last = legs[0], legs[-1]
    fel_hu, le_hu = first['fel_hu'], last['le_hu']
    viszonylat = f"{first['fel']}  →  {last['le']}"
    tipus_str = first['tipus']
    expected = _classify_fuvarfeladat_type_expected(tipus_str)

    def _res(hiba, javaslat, implies_complete=False, window_truncated=False,
             window_start_truncated=False):
        return {
            'torzs': torzs,
            'tipus': tipus_str,
            'viszonylat': viszonylat,
            'fuvarszamok': ', '.join(l['fsz'] for l in legs),
            'jaratszamok': ', '.join(dict.fromkeys(l['jsz'] for l in legs)),
            'hiba': hiba,
            'javaslat': javaslat,
            'implies_complete': implies_complete,
            'window_truncated': window_truncated,
            'window_start_truncated': window_start_truncated,
        }

    # A törzs utolsó aktivitása + időszak-vég közelség (v4.6): ha a törzs a
    # feltöltött adatablak utolsó napjaiban szakad félbe, a hiányzó nemzetközi /
    # hazatérő részfeladat valószínűleg az időszak UTÁN van → nem hiba.
    _dates = [d for l in legs for d in (l['_fel_raw'], l['_le_raw']) if d is not None]
    last_activity = max(_dates) if _dates else None
    first_activity = min(_dates) if _dates else None
    # v4.7: az adatablak ELEJÉN csonkolt törzs (a v4.6-os ablak-vég logika tükörképe)
    near_window_start = (
        window_start is not None and pd.notna(window_start)
        and first_activity is not None
        and (pd.Timestamp(first_activity) - pd.Timestamp(window_start)).days
        <= WINDOW_TRUNCATION_DAYS
    )
    near_window_end = (
        window_end is not None and pd.notna(window_end)
        and last_activity is not None
        and (pd.Timestamp(window_end) - pd.Timestamp(last_activity)).days
        <= WINDOW_TRUNCATION_DAYS
    )

    def _truncated_start_res(reszlet):
        return _res(
            f"{reszlet} A törzs első aktivitása ({first_activity:%Y-%m-%d}) a feltöltött "
            f"adatablak első napjaira esik (ablak eleje: {pd.Timestamp(window_start):%Y-%m-%d}) "
            f"— a kör valószínűleg a lekérdezett időszak ELŐTT kezdődött. Ez nem rögzítési hiba.",
            "Futtasd újra az előző időszak fuvarnaplójával együtt, akkor a kör teljes lesz. "
            "Addig a kör a pénzügyi kimutatásokból kizárva.",
            window_truncated=True,
            window_start_truncated=True,
        )

    def _truncated_res(reszlet):
        return _res(
            f"{reszlet} A törzs utolsó aktivitása ({last_activity:%Y-%m-%d}) a feltöltött "
            f"adatablak utolsó napjaira esik (ablak vége: {pd.Timestamp(window_end):%Y-%m-%d}) "
            f"— a kör valószínűleg a lekérdezett időszak UTÁN fejeződik be. Ez nem rögzítési hiba.",
            "Futtasd újra a következő időszak fuvarnaplójával együtt, akkor a kör teljes lesz. "
            "Addig a kör a pénzügyi kimutatásokból kizárva.",
            window_truncated=True,
        )

    # Közbenső mintázatok
    max_dij = max(l['dij'] for l in legs)
    # Hurok-láb: felvétel == leadás (és nem 0 díjas apró korrekciós sor).
    # Tipikusan a nemzetközi oda-vissza út, ahol a külföldi állomások "eltűntek",
    # mert a hazafelé felrakó lerakóként lett rögzítve.
    hurok_legs = [l for l in legs
                  if l['fel'] and l['fel'] == l['le'] and l['dij'] >= max_dij * 0.5 and l['dij'] > 0]
    # Látható külföldi lerakó valamelyik közbenső lábon
    kozbenso_kulfoldi_le = [l for l in legs[:-1] if not l['le_hu']]
    kozbenso_kulfoldi_fel = [l for l in legs[1:] if not l['fel_hu']]

    if expected == 'EXPORT':
        if fel_hu and not le_hu:
            return None  # szabályos export
        if fel_hu and le_hu:
            if kozbenso_kulfoldi_le:
                k = kozbenso_kulfoldi_le[0]
                idx = legs.index(k)
                next_leg = legs[idx + 1] if idx + 1 < len(legs) else None
                if next_leg is not None and next_leg['fel'] == k['le']:
                    return _res(
                        f"A feladat kimegy külföldre ({k['fsz']}: lerakó {k['le']}), majd onnan "
                        f"felrakóval hazatér és itthon lerak ({last['fsz']}: {last['le']}). "
                        f"Ez a viszonylat teljes kör, nem Export.",
                        "Fuvarfeladat típusa javítandó: Export → EU körfuvar.",
                        implies_complete=True,
                    )
                return _res(
                    f"A feladat kimegy külföldre ({k['fsz']}: lerakó {k['le']}), és itthon zár "
                    f"({last['fsz']}: {last['le']}), de a hazafelé viszonylaton nincs rögzített "
                    f"külföldi FELRAKÓ. Komplett körnél a hazaútra mindenképp fel kell rakni — "
                    f"feltehetően állomás-típus hiba (lerakó lett rögzítve felrakó helyett).",
                    "Fuvarfeladat típusa javítandó: Export → EU körfuvar, ÉS a hazafelé szakasz "
                    "külföldi állomása felrakóra javítandó.",
                    implies_complete=True,
                )
            if hurok_legs:
                h = hurok_legs[0]
                return _res(
                    f"A törzs minden állomása HU, de a(z) {h['fsz']} részfeladat hurok-viszonylat "
                    f"({h['fel']} → ugyanoda, díj {h['dij']:.0f} EUR) — ez a nemzetközi oda-vissza út. "
                    f"A külföldi állomás azért nem látszik, mert a kinti LERAKÓ után a hazafelé "
                    f"FELRAKÓ is lerakóként lett rögzítve, így az 'első felrakó / utolsó lerakó' "
                    f"nézetben mindkettő kiesik. Állomás-típus hiba a forrásrendszerben.",
                    "Fuvarfeladat típusa javítandó: Export → EU körfuvar, ÉS a hazafelé induló "
                    "külföldi állomás típusa lerakóról felrakóra javítandó.",
                    implies_complete=True,
                )
            if near_window_end:
                return _truncated_res(
                    "A törzs minden rögzített állomása HU, a nemzetközi részfeladat még hiányzik."
                )
            return _res(
                f"Export típusnál HU felrakó és nem-HU lerakó kellene, itt minden állomás HU. "
                f"Nemzetközi részfeladat nem található a törzsben — vagy nincs rögzítve, vagy a "
                f"lekérdezett időszakon kívül esik (adatablak-csonkolás).",
                "Ellenőrizd a forrásrendszerben: van-e a törzsnek nemzetközi részfeladata. "
                "Ha nincs, a típus javítandó (pl. Belföldi fuvar); ha van, az időablak bővítendő.",
            )
        if (not fel_hu) and le_hu:
            return _res(
                f"A viszonylat külföldről HU-ba tart ({viszonylat}) — ez Import irány, nem Export.",
                "Fuvarfeladat típusa javítandó: Export → Import.",
            )
        return _res(
            f"Sem a felrakó, sem a lerakó nem HU ({viszonylat}).",
            "Fuvarfeladat típusa javítandó: Export → Harmadik országba szállítás.",
        )

    if expected == 'IMPORT':
        if (not fel_hu) and le_hu:
            return None  # szabályos import
        if (not fel_hu) and (not le_hu):
            return _res(
                f"A törzs külföldről indul, de a záró részfeladat ({last['fsz']}) külföldre visz "
                f"({last['fel']} → {last['le']}). A záró láb nem Import viszonylat.",
                f"A(z) {last['fsz']} részfeladat típusa javítandó: Import → Export vagy "
                f"Harmadik országba szállítás (viszonylattól függően).",
            )
        if fel_hu and le_hu:
            if kozbenso_kulfoldi_fel:
                k = kozbenso_kulfoldi_fel[0]
                return _res(
                    f"A törzs HU-ban indul és zár, de közben külföldi felrakó van "
                    f"({k['fsz']}: {k['fel']}). Ez teljes kör mintázat, nem Import.",
                    "Fuvarfeladat típusa javítandó: Import → EU körfuvar.",
                    implies_complete=True,
                )
            if hurok_legs:
                h = hurok_legs[0]
                return _res(
                    f"A törzs minden állomása HU, de a(z) {h['fsz']} részfeladat hurok-viszonylat "
                    f"({h['fel']} → ugyanoda, díj {h['dij']:.0f} EUR) — ez a nemzetközi oda-vissza út, "
                    f"a külföldi állomások állomás-típus hiba miatt nem látszanak.",
                    "Fuvarfeladat típusa javítandó: Import → EU körfuvar, ÉS a külföldi állomások "
                    "fel-/lerakó típusa ellenőrizendő.",
                    implies_complete=True,
                )
            if near_window_start:
                return _truncated_start_res(
                    "Import típusnál a nemzetközi (külföld → HU) részfeladat hiányzik, "
                    "csak a belföldi kiosztó láb van meg."
                )
            if near_window_end:
                return _truncated_res(
                    "A törzs minden rögzített állomása HU, a nemzetközi részfeladat még hiányzik."
                )
            return _res(
                f"Import típusnál nem-HU felrakó és HU lerakó kellene, itt minden állomás HU. "
                f"Nemzetközi részfeladat nem található — vagy nincs rögzítve, vagy a lekérdezett "
                f"időszakon kívül esik.",
                "Ellenőrizd a forrásrendszerben a törzs nemzetközi részfeladatát; ha nincs, a "
                "típus javítandó.",
            )
        return _res(
            f"A viszonylat HU-ból külföldre tart ({viszonylat}) — ez Export irány, nem Import.",
            "Fuvarfeladat típusa javítandó: Import → Export.",
        )

    if expected == 'EU_KOR':
        if fel_hu and le_hu:
            return None  # szabályos EU körfuvar
        if fel_hu and not le_hu:
            if near_window_end:
                return _truncated_res(
                    f"Az utolsó rögzített lerakó külföldön van ({last['fsz']}: {last['le']}), "
                    f"a hazatérő szakasz még hiányzik."
                )
            return _res(
                f"EU körfuvarnak HU-ban kell zárnia, de az utolsó rögzített lerakó külföldön van "
                f"({last['fsz']}: {last['le']}). A hazatérő szakasz hiányzik: vagy nincs rögzítve "
                f"a hazafelé felrakó + itthoni lerakó, vagy a hazatérő részfeladat az időszakon "
                f"kívül esik.",
                "A forrásrendszerben rögzítendő a hazatérő szakasz (külföldi FELRAKÓ + HU lerakó), "
                "vagy ha a fuvar tényleg kint zárt, a típus javítandó: EU körfuvar → Export.",
            )
        return _res(
            f"EU körfuvar típusnál HU indulás és HU érkezés kellene, itt a viszonylat {viszonylat}.",
            "Ellenőrizd a felrakó állomás rögzítését, vagy a típus javítandó.",
        )

    # SEMLEGES / ISMERETLEN → nem validáljuk
    return None


def _sort_legs_by_reszfeladat(group: pd.DataFrame) -> pd.DataFrame:
    """Lábak sorrendje a Fuvarszám részfeladat-sorszáma szerint (a v4.6-os viselkedés)."""
    g = group.copy()
    if '_reszfeladat' not in g.columns:
        g['_reszfeladat'] = g['Fuvarszám'].map(_reszfeladat_of)
    if g['_reszfeladat'].notna().any():
        return g.sort_values(['_reszfeladat', 'Első Felvételi állomás időkapu (dátum)'],
                             na_position='last')
    return g.sort_values('Első Felvételi állomás időkapu (dátum)')


# A sorrend-ellentmondás miatt "megmentett" törzsek: a részfeladat-sorszám szerinti
# lánc típushibásnak látszik, az időrend szerinti viszont konzisztens (v4.7).
_ORDER_RESOLVED_TORZSEK = {}


def analyze_torzs_type(torzs: str, torzs_group: pd.DataFrame, window_end=None,
                       window_start=None):
    """Típus-/állomás-validáció sorrend-robusztusan (v4.7).

    A validáció alapértelmezésben a részfeladat-sorszám szerinti láncot vizsgálja.
    Az ügyfél ellenőrzése kimutatta, hogy a sorszámok egy részénél a 2. és 3.
    részfeladat fel van cserélve, miközben az időkapuk helyesek — ilyenkor a
    sorszám szerinti lánc hamis "hibás fuvarfeladat típus" riasztást ad.

    Ezért: ha a sorszám szerinti lánc hibát jelez, DE a két sorrend eltér és az
    időrend szerinti lánc hibátlan, akkor nem típushibáról, hanem rögzítési
    sorrend-hibáról van szó → nem adunk típushibát, a törzs a
    "Részfeladat-sorrend eltérés" riportba kerül.
    """
    _ckey = str(torzs)
    if _ckey in _TORZS_ANALYSIS_CACHE:
        return _TORZS_ANALYSIS_CACHE[_ckey]

    res = _analyze_torzs_type_ordered(torzs, torzs_group, window_end=window_end,
                                      order='reszfeladat', window_start=window_start)
    if res is None:
        _TORZS_ANALYSIS_CACHE[_ckey] = None
        return None

    # Alternatív, ugyanennyire hihető lábsorrendek. Ha BÁRMELYIK konzisztens
    # láncot ad, a rögzítési sorrend a bizonytalan, nem a fuvarfeladat típusa.
    _recs = _leg_records(torzs, torzs_group)
    base_seq = [l['fsz'] for l in _order_records(_recs, 'reszfeladat')]
    for alt_order in ('chrono', 'leadas'):
        if [l['fsz'] for l in _order_records(_recs, alt_order)] == base_seq:
            continue
        alt = _analyze_torzs_type_ordered(torzs, torzs_group, window_end=window_end,
                                          order=alt_order, window_start=window_start)
        if alt is None:
            _ORDER_RESOLVED_TORZSEK[str(torzs)] = res.get('hiba', '')
            _TORZS_ANALYSIS_CACHE[_ckey] = None
            return None

    # Marad a hiba – de ha dátum-anomália is van a törzsben (leadás < felvétel),
    # azt jelezzük, mert a hibás időkapu maga is okozhatja a téves láncot.
    if _van_datum_anomalia(torzs_group):
        res = dict(res)
        res['hiba'] = (res['hiba'] + ' ⚠️ FIGYELEM: a törzsben dátum-anomália is van '
                       '(leadási időkapu korábbi a felvételinél) — a hibajelzés ebből '
                       'is eredhet, előbb az időkapukat érdemes ellenőrizni.')
    _TORZS_ANALYSIS_CACHE[_ckey] = res
    return res


def validate_torzs_type(torzs: str, torzs_group: pd.DataFrame, window_end=None,
                        window_start=None):
    """Egy fuvarszám-törzsre visszaadja a fuvarfeladat típus validációs hibáit
    (szöveges lista a kör-magyarázathoz). A részletes elemzést az
    analyze_torzs_type végzi.

    v4.6: az időszak-vég miatt csonkolt törzs (window_truncated) NEM hiba —
    arra a kör-szintű kék 'időszak után záródó' jelölés hívja fel a figyelmet."""
    res = analyze_torzs_type(torzs, torzs_group, window_end=window_end,
                             window_start=window_start)
    if res is None or res.get('window_truncated'):
        return []
    return [
        f"Hibás fuvarfeladat típus – törzs {torzs} ({res['tipus']}), viszonylat: "
        f"{res['viszonylat']}. {res['hiba']} 👉 {res['javaslat']}"
    ]


# ---------------------------------------------------------------------------
# A kör tartalma alapján előálló magyarázat + szín (részletes)
# ---------------------------------------------------------------------------
def _format_leg_step(idx, fsz, jsz, ir, fel_cim, le_cim):
    """Egy sor a kör lánc-leírásából."""
    return f"{idx}. {fsz} ({jsz}, {ir}): {fel_cim}  →  {le_cim}"


def _build_kor_content_explanation(legs_ordered, ures_visszafutas_gyanu=False,
                                   implies_complete=False,
                                   harmadik_orszagos_zaras=False):
    """A kör tartalma alapján visszaadja a magyarázatot és a színt.
    legs_ordered: [(fuvarszám, járatszám, irány, felvétel_cím, leadás_cím), ...] időrendben.

    v4.1: a 'korfuvar' irányú láb (HU→HU EU körfuvar) kifelé ÉS befelé szakasznak is
    számít – önmagában teljes kört ad. Kifelé-only körnél az üres visszafutás gyanú
    külön megjegyzést kap.
    v4.2: implies_complete=True, ha a típus-elemzés szerint a kör a valóságban
    lezárult (EU körfuvar mintázat rossz típussal / állomás-típus hibával rögzítve).
    """
    has_korfuvar = any(l[2] == 'korfuvar' for l in legs_ordered)
    has_kifele = any(l[2].startswith('kifelé') for l in legs_ordered) or has_korfuvar or implies_complete
    has_befele = any(l[2].startswith('befelé') for l in legs_ordered) or has_korfuvar or implies_complete
    has_semleges = any(l[2] == 'semleges' for l in legs_ordered)

    if has_kifele and has_befele:
        msg = 'Teljes kör: kifelé és befelé szakasz is lezárult a körben.'
        if has_korfuvar:
            msg = ('Teljes kör: EU körfuvar (oda-vissza) szakaszt tartalmaz, '
                   'amely önmagában lezárja a kört.')
        if implies_complete and not (any(l[2].startswith('befelé') for l in legs_ordered) or has_korfuvar):
            msg = ('Teljes kör: a rögzítés alapján EU körfuvar mintázat — a jármű kiment és '
                   'hazatért, de a fuvarfeladat típusa/állomásai hibásan lettek rögzítve '
                   '(részletek a típushibánál).')
        return msg, 'background-color: lightgreen'

    # Részletes láncleírás a részleges körhöz
    steps = [_format_leg_step(i, l[0], l[1], l[2], l[3], l[4]) for i, l in enumerate(legs_ordered, 1)]
    lanc = ' || '.join(steps)

    if not has_kifele and has_befele and not has_semleges:
        first = legs_ordered[0]
        msg = (
            f"Részleges kör: nincs export (kifelé) nyitás — az első fuvarfeladat is "
            f"import (befelé) volt: {first[0]} ({first[1]}), viszonylat: {first[3]}  →  {first[4]}. "
            f"A körben csak import típusú járatok vannak. Lánc: {lanc}"
        )
        return msg, 'background-color: orange'

    # v4.7: harmadik országos zárás – az ügyfél visszaigazolta, hogy ha a kör záró
    # fuvarfeladata "Harmadik országba szállítás" típusú és nincs magyar állomása,
    # az NEM hiányzó import, hanem szabályos, külföldön záruló harmadik országos
    # szolgáltatás. Ilyenkor nem adunk hibát.
    if harmadik_orszagos_zaras and has_kifele and not has_befele:
        last = legs_ordered[-1]
        msg = (
            f"Harmadik országos zárás (nem hiba): a kör záró fuvarfeladata "
            f"'Harmadik országba szállítás' típusú, magyar állomás nélkül: "
            f"{last[0]} ({last[1]}), viszonylat: {last[3]}  →  {last[4]}. "
            f"A hiányzó záró import ilyenkor nem rögzítési hiba. Lánc: {lanc}"
        )
        return msg, 'background-color: lightgreen'

    if has_kifele and not has_befele and not has_semleges:
        last = legs_ordered[-1]
        msg = (
            f"Részleges kör: hiányzik a záró import (befelé) fuvar — az utolsó fuvarfeladat "
            f"is export (kifelé) volt: {last[0]} ({last[1]}), viszonylat: {last[3]}  →  {last[4]}. "
            f"A körben csak export típusú járatok vannak."
        )
        if ures_visszafutas_gyanu:
            msg += (" A vontatmány következő fuvarfeladata is kifelé irányú → "
                    "valószínű üres visszafutás (nincs rögzített import a visszaútra).")
        msg += f" Lánc: {lanc}"
        return msg, 'background-color: orange'

    if has_kifele and has_semleges and not has_befele:
        msg = (
            f"Részleges kör: van export (kifelé) nyitás és semleges (harmadik országos) szakasz is, "
            f"de hiányzik a záró import (befelé) fuvar. Lánc: {lanc}"
        )
        return msg, 'background-color: orange'

    if not has_kifele and has_semleges and has_befele:
        msg = (
            f"Részleges kör: van semleges és import szakasz is, de hiányzik a nyitó "
            f"export (kifelé) fuvar — az első feladat sem export. Lánc: {lanc}"
        )
        return msg, 'background-color: orange'

    if not has_kifele and has_semleges and not has_befele:
        msg = (
            f"Részleges kör: csak semleges (harmadik országos) szakasz(ok) vannak, "
            f"sem export, sem import fuvar nem tartozik hozzá. Lánc: {lanc}"
        )
        return msg, 'background-color: orange'

    msg = f"Részleges / ismeretlen kör (irány nem sorolható be). Lánc: {lanc}"
    return msg, 'background-color: orange'


def _build_vontatmany_change_explanation(torzs: str, torzs_history: list):
    """torzs_history: [{'fuvarszám','járatszám','vontatmány'}, ...] a törzs
    időrendben rendezett részfeladatai (a részfeladat sorszáma szerint, ha van, egyébként idő szerint).
    Ha egynél több vontatmány található, elmagyarázza, hol váltott.
    Visszatér: string vagy None."""
    if not torzs_history or len(torzs_history) < 2:
        return None
    egyedi_v = list(dict.fromkeys(str(h.get('vontatmány', '')) for h in torzs_history if h.get('vontatmány') is not None))
    if len(egyedi_v) <= 1:
        return None
    for i in range(1, len(torzs_history)):
        prev = torzs_history[i - 1]
        cur = torzs_history[i]
        if str(cur.get('vontatmány', '')) != str(prev.get('vontatmány', '')):
            return (
                f"a {torzs} fuvarszám törzs több vontatmányon fut. "
                f"A(z) {prev.get('fuvarszám','?')} feladat (járat {prev.get('járatszám','?')}) "
                f"a(z) '{prev.get('vontatmány','?')}' vontatmánnyal indult, majd a következő "
                f"{cur.get('fuvarszám','?')} feladatnál (járat {cur.get('járatszám','?')}) "
                f"'{cur.get('vontatmány','?')}'-ra változott. "
                f"Érintett vontatmányok összesen: {', '.join(egyedi_v)}."
            )
    return (
        f"a {torzs} fuvarszám törzs több vontatmányon fut: "
        f"{', '.join(egyedi_v)}."
    )


def build_full_kor_explanation(legs_ordered, torzsek_a_korben, torzs_history_map,
                               torzs_group_map, ures_visszafutas_gyanu=False,
                               implies_complete=False, idoszak_utan_zarodo=False,
                               window_end=None, harmadik_orszagos_zaras=False,
                               idoszak_elott_kezdodo=False, window_start=None):
    """Összeállítja a kör teljes magyarázatát több szempont figyelembevételével:
      1) Vontatmány váltás minden érintett törzsre
         v4.1: ha a törzs MINDEN részfeladata ebben a körben van (a törzs-alapú
         kör-összefűzés után ez a normál eset), a vontatmány-váltás elő-fuvar /
         pótkocsi-csere INFÓ, nem hiba. Csak akkor piros, ha a törzs tényleg
         több körre esett szét.
      2) Fuvarfeladat típus validáció minden érintett törzsre
      3) A kör tartalma (teljes / részleges) – részletes láncleírással
    Szín prioritás: szétesett törzs (piros) > típus/tartalmi hiba (narancs) > zöld.
    """
    parts = []
    has_vontatmany_hiba = False
    has_tipus_hiba = False

    kor_fuvarszamok = {l[0] for l in legs_ordered}

    # 0) Időszak után záródó kör (v4.6) — a legelső helyen jelezzük
    if idoszak_utan_zarodo:
        parts.append(
            '⏭️ Időszak után záródó kör (valószínű): a kör a feltöltött adatablak utolsó '
            'napjaiban szakad félbe — a záró szakasz / hiányzó részfeladat a következő '
            'időszakban várható. Ez NEM rögzítési hiba; a kör a pénzügyi kimutatásokból '
            'kizárva. A következő időszak fuvarnaplójával együtt futtatva teljes lesz.'
        )

    # 0/b) Időszak ELŐTT kezdődő kör (v4.7)
    if idoszak_elott_kezdodo:
        parts.append(
            '⏮️ Időszak előtt kezdődő kör (valószínű): a kör nyitó (export) szakasza a '
            'feltöltött adatablak kezdete ELŐTT van — itt csak a hazafelé tartó / belföldi '
            'kiosztó rész látszik. Ez NEM rögzítési hiba; a kör a pénzügyi kimutatásokból '
            'kizárva. Az előző időszak fuvarnaplójával együtt futtatva teljes lesz.'
        )

    # 1) Vontatmány váltás
    for torzs in torzsek_a_korben:
        history = torzs_history_map.get(torzs, [])
        msg = _build_vontatmany_change_explanation(torzs, history)
        if msg:
            torzs_all_in = all(str(h.get('fuvarszám')) in kor_fuvarszamok for h in history)
            if torzs_all_in:
                parts.append(
                    'ℹ️ Elő-fuvar / pótkocsi-csere (nem hiba): ' + msg +
                    ' A törzs összes részfeladata ebben a körben van.'
                )
            else:
                parts.append('Változó vontatmány hiba – ' + msg +
                             ' A törzs részfeladatai több körre estek szét.')
                has_vontatmany_hiba = True

    # 2) Fuvarfeladat típus validáció
    for torzs in torzsek_a_korben:
        grp = torzs_group_map.get(torzs)
        if grp is None:
            continue
        errs = validate_torzs_type(torzs, grp, window_end=window_end,
                                   window_start=window_start)
        for e in errs:
            parts.append(e)
            has_tipus_hiba = True

    # 2/b) Részfeladat-sorrend eltérés (v4.7) – NEM hiba: a kör legenerálva, csak jelzés
    for torzs in torzsek_a_korben:
        if str(torzs) in _ORDER_RESOLVED_TORZSEK:
            parts.append(
                f'ℹ️ Részfeladat-sorrend eltérés (nem hiba): a(z) {torzs} törzsnél a '
                f'részfeladat-sorszámok ellentmondanak az időkapuknak. Az időkapuk szerinti '
                f'lánc konzisztens, ezért a kör hibátlanul legenerálva — a sorszámozás a '
                f'forrásrendszerben javítandó (lásd a "Részfeladat-sorrend eltérés" munkalapot).'
            )

    # 3) Kör tartalom
    content_msg, content_color = _build_kor_content_explanation(
        legs_ordered, ures_visszafutas_gyanu=ures_visszafutas_gyanu,
        implies_complete=implies_complete,
        harmadik_orszagos_zaras=harmadik_orszagos_zaras)
    parts.append(content_msg)

    combined = '  ///  '.join(parts)

    # Szín prioritás: szétesett törzs > időszak után záródó (kék) > típus hiba > tartalom
    if has_vontatmany_hiba:
        color = 'background-color: lightcoral'
    elif idoszak_utan_zarodo or idoszak_elott_kezdodo:
        color = 'background-color: lightblue'
    elif has_tipus_hiba:
        color = 'background-color: orange'
    else:
        color = content_color

    return combined, color


# ---------------------------------------------------------------------------
# Kör-építés + törzs-alapú kör-összefűzés (v4.1)
# ---------------------------------------------------------------------------
def build_korfuvarok(df: pd.DataFrame):
    """Körök építése vontatmányonként, majd a közös fuvarszám-törzsön osztozó
    körök összefűzése (union-find).

    Az elő-fuvar minta miatt (a -1 belföldi felfutás shuttle vontatmánnyal megy a
    bábolnai hubra, a -2 nemzetközi láb a vonali vontatmánnyal) egy törzs lábai
    különböző vontatmány-csoportokba kerülhetnek. Az összefűzés ezeket egyetlen
    körré egyesíti, így a törzs sosem esik több körre.

    Visszatér: [(kor_id, vontatmany_label, legs_list), ...]
    """
    korfuvarok = []
    global_kor_id = 0

    for vontatmany, grp in df.groupby('Vontatmány'):
        grp_sorted = grp.sort_values([
            'Utolsó Leadási állomás időkapu (dátum)',
            'Első Felvételi állomás időkapu (dátum)',
        ])

        current_kor_legs = []
        current_fuv_torzsek = set()
        current_jaratszamok = set()

        for _, row in grp_sorted.iterrows():
            f_szam = str(row['Fuvarszám'])
            j_szam = str(row['Járatszám'])
            f_torzs = _torzs_of(f_szam)
            irany = row['Irány']

            if not current_kor_legs:
                global_kor_id += 1
                current_kor_legs = [row]
                current_fuv_torzsek = {f_torzs}
                current_jaratszamok = {j_szam}
                continue

            prev_irany = current_kor_legs[-1]['Irány']

            kapcsolodik_szam = (
                f_torzs in current_fuv_torzsek or j_szam in current_jaratszamok
            )

            irany_osszetartozo = False
            if prev_irany.startswith('kifelé') and irany in (
                    'semleges', 'befelé-nemzetközi', 'befelé-belföldi', 'korfuvar'):
                irany_osszetartozo = True
            if prev_irany == 'semleges' and irany in (
                    'semleges', 'befelé-nemzetközi', 'befelé-belföldi', 'korfuvar'):
                irany_osszetartozo = True

            if kapcsolodik_szam or irany_osszetartozo:
                current_kor_legs.append(row)
                current_fuv_torzsek.add(f_torzs)
                current_jaratszamok.add(j_szam)
                continue

            korfuvarok.append((global_kor_id, vontatmany, current_kor_legs))
            global_kor_id += 1
            current_kor_legs = [row]
            current_fuv_torzsek = {f_torzs}
            current_jaratszamok = {j_szam}

        if current_kor_legs:
            korfuvarok.append((global_kor_id, vontatmany, current_kor_legs))

    # --- Körök összefűzése: közös fuvarszám-törzsön osztozó körök egyesítése ---
    ring_torzsek = [
        {_torzs_of(str(r['Fuvarszám'])) for r in legs}
        for _, _, legs in korfuvarok
    ]
    parent = list(range(len(korfuvarok)))

    def _find(x):
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x

    torzs_to_ring = {}
    for i, ts in enumerate(ring_torzsek):
        for t in ts:
            if t in torzs_to_ring:
                ra, rb = _find(torzs_to_ring[t]), _find(i)
                if ra != rb:
                    parent[rb] = ra
            else:
                torzs_to_ring[t] = i

    merged = {}
    for i in range(len(korfuvarok)):
        merged.setdefault(_find(i), []).append(korfuvarok[i])

    korfuvarok_merged = []
    for grp in merged.values():
        all_legs = [r for _, _, legs in grp for r in legs]
        vontok = ' + '.join(sorted({str(v) for _, v, _ in grp}))
        korfuvarok_merged.append((grp[0][0], vontok, all_legs))

    # Kör ID-k újraszámozása determinisztikus sorrendben (első láb felvételi ideje)
    def _first_pickup(item):
        dts = [r['Első Felvételi állomás időkapu (dátum)'] for r in item[2]
               if pd.notna(r['Első Felvételi állomás időkapu (dátum)'])]
        return min(dts) if dts else pd.Timestamp.max

    korfuvarok_merged.sort(key=_first_pickup)
    return [(i + 1, vont, legs) for i, (_, vont, legs) in enumerate(korfuvarok_merged)]


def detect_ures_visszafutas(df: pd.DataFrame, legs_df: pd.DataFrame, kor_veg) -> bool:
    """Kifelé-only körnél: a kör vontatmányán a kör vége UTÁNI következő láb
    megint kifelé irányú-e (→ valószínű üres visszafutás)."""
    if pd.isna(kor_veg):
        return False
    vontok = set(legs_df['Vontatmány'].astype(str))
    after = df[
        df['Vontatmány'].astype(str).isin(vontok)
        & (df['Utolsó Leadási állomás időkapu (dátum)'] > kor_veg)
    ]
    if after.empty:
        return False
    nxt = after.sort_values('Utolsó Leadási állomás időkapu (dátum)').iloc[0]
    return str(nxt['Irány']).startswith('kifelé')


def generate_result_df(df: pd.DataFrame) -> pd.DataFrame:
    """A teljes kör-generálási pipeline a fuvarnaplóból (hónap-szűrés NÉLKÜL).

    Elvárás: a df-ben az időkapu oszlopok datetime-ok és az 'Irány' oszlop kitöltött.
    """
    # Cím-gazetteer építése a teljes adathalmazból (v4.7) – a prefix nélküli
    # címek országa a prefixes címekből tanult (irányítószám, város) párokból jön.
    build_address_gazetteer(df)
    _reset_torzs_caches()

    # Előindexelés törzsenként (vontatmány-váltás leíráshoz és típus-validációhoz)
    tmp_torzs = df['Fuvarszám'].astype(str).map(_torzs_of)
    df_indexed = df.assign(_torzs=tmp_torzs)

    torzs_history_map = {}
    for torzs, grp in df_indexed.groupby('_torzs'):
        torzs_history_map[torzs] = [
            {
                'fuvarszám': l['fsz'],
                'járatszám': l['jsz'],
                'vontatmány': l['vontatmany'],
            }
            for l in _order_records(_leg_records(torzs, grp), 'reszfeladat')
        ]

    torzs_group_map = {
        torzs: grp.drop(columns=['_torzs'], errors='ignore')
        for torzs, grp in df_indexed.groupby('_torzs')
    }

    # A feltöltött adatablak vége (időszak-vég csonkolás felismeréséhez, v4.6)
    window_end = df['Utolsó Leadási állomás időkapu (dátum)'].max()
    # v4.7: az adatablak ELEJE (a v4.6-os ablak-vég logika tükörképe)
    window_start = df['Utolsó Leadási állomás időkapu (dátum)'].min()

    # Törzs-szintű típus-elemzés (1x, cache-elve)
    torzs_analysis_map = {
        torzs: analyze_torzs_type(torzs, grp, window_end=window_end,
                                  window_start=window_start)
        for torzs, grp in torzs_group_map.items()
    }

    korfuvarok = build_korfuvarok(df)

    output_rows = []
    for kor_id, vontatmany, legs in korfuvarok:
        legs_df = pd.DataFrame(legs)

        total_dij = legs_df['Díj részarány (EUR)'].sum() if 'Díj részarány (EUR)' in legs_df.columns else 0

        all_vontatok = ' | '.join(
            legs_df['Vontató'].astype(str).dropna().unique().tolist()
        )
        jaratszamok_lista = (
            legs_df['Járatszám'].astype(str).str.strip().dropna().unique().tolist()
        )
        all_jaratszamok = ' | '.join(jaratszamok_lista)

        kifele_legs = legs_df[legs_df['Irány'].str.startswith('kifelé')]
        befele_legs = legs_df[legs_df['Irány'].str.startswith('befelé')]
        semleges_legs = legs_df[legs_df['Irány'] == 'semleges']
        korfuvar_legs = legs_df[legs_df['Irány'] == 'korfuvar']

        kif_kezd_ido, kif_kezd_cim, kif_zar_ido, kif_zar_cim = get_interval_with_addresses(kifele_legs)
        sem_kezd_ido, sem_kezd_cim, sem_zar_ido, sem_zar_cim = get_interval_with_addresses(semleges_legs)
        bef_kezd_ido, bef_kezd_cim, bef_zar_ido, bef_zar_cim = get_interval_with_addresses(befele_legs)
        krf_kezd_ido, _krf_kezd_cim, krf_zar_ido, _krf_zar_cim = get_interval_with_addresses(korfuvar_legs)

        torzsek_a_korben = list(dict.fromkeys(
            _torzs_of(f) for f in legs_df['Fuvarszám'].astype(str)))

        # A típus-elemzés szerint a kör a valóságban lezárult-e (pl. Wolfsburg minta:
        # Export típus, de hurok-lábbal hazatért — csak a rögzítés hibás)
        implies_complete = any(
            (torzs_analysis_map.get(t) or {}).get('implies_complete', False)
            for t in torzsek_a_korben
        )

        has_korfuvar = not korfuvar_legs.empty
        has_kifele = (not kifele_legs.empty) or has_korfuvar or implies_complete
        has_befele = (not befele_legs.empty) or has_korfuvar or implies_complete
        has_semleges = not semleges_legs.empty

        # Kör kezdete: kifelé → korfuvar → semleges → befelé (fallback lánc)
        kor_kezd = kif_kezd_ido
        for cand in (krf_kezd_ido, sem_kezd_ido, bef_kezd_ido):
            if pd.isna(kor_kezd):
                kor_kezd = cand
        # Kör vége: befelé → korfuvar → semleges → kifelé
        kor_veg = bef_zar_ido
        for cand in (krf_zar_ido, sem_zar_ido, kif_zar_ido):
            if pd.isna(kor_veg):
                kor_veg = cand

        # Időszak után záródó kör (v4.6): vagy ablak-vég miatt csonkolt törzs van
        # benne, vagy import-zárás nélküli kör, amelynek vége az adatablak utolsó
        # napjaira esik (a záró import a következő időszakban várható)
        idoszak_utan_zarodo = any(
            (torzs_analysis_map.get(t) or {}).get('window_truncated', False)
            for t in torzsek_a_korben
        )
        if (not idoszak_utan_zarodo and has_kifele and not has_befele
                and pd.notna(kor_veg) and pd.notna(window_end)
                and (pd.Timestamp(window_end) - pd.Timestamp(kor_veg)).days
                <= WINDOW_TRUNCATION_DAYS):
            idoszak_utan_zarodo = True

        # Időszak ELŐTT kezdődő kör (v4.7): export-nyitás nélküli (csak import)
        # kör, amelynek kezdete az adatablak első napjaira esik — a nyitó export
        # a lekérdezett időszak előtt van, tehát nem rögzítési hiba.
        idoszak_elott_kezdodo = (
            not idoszak_utan_zarodo and has_befele and not has_kifele
            and pd.notna(kor_kezd) and pd.notna(window_start)
            and (pd.Timestamp(kor_kezd) - pd.Timestamp(window_start)).days
            <= WINDOW_TRUNCATION_DAYS
        )
        if not idoszak_elott_kezdodo and not idoszak_utan_zarodo:
            idoszak_elott_kezdodo = any(
                (torzs_analysis_map.get(t) or {}).get('window_truncated', False)
                and (torzs_analysis_map.get(t) or {}).get('window_start_truncated', False)
                for t in torzsek_a_korben
            )

        # Időrendben rendezett részfeladat lista a magyarázathoz
        legs_sorted_for_expl = legs_df.sort_values([
            'Első Felvételi állomás időkapu (dátum)',
            'Utolsó Leadási állomás időkapu (dátum)',
        ], na_position='last')
        legs_ordered_tuples = [
            (
                str(r['Fuvarszám']),
                str(r['Járatszám']),
                str(r['Irány']),
                str(r.get('Első Felvételi állomás cím', '') or ''),
                str(r.get('Utolsó Leadási állomás cím', '') or ''),
            )
            for _, r in legs_sorted_for_expl.iterrows()
        ]

        # Üres visszafutás gyanú (csak kifelé-only körnél érdekes)
        ures_gyanu = False
        if has_kifele and not has_befele:
            ures_gyanu = detect_ures_visszafutas(df, legs_df, kor_veg)

        # Ország-reláció: a lábak időrendi lánca alapján (egymást követő azonos
        # országok összevonva), pl. 'HU→DE→HU'. Célország = az első nem-HU ország.
        chain = []
        for l in legs_ordered_tuples:
            for cim in (l[3], l[4]):
                cc = country_of(cim)
                if not chain or chain[-1] != cc:
                    chain.append(cc)
        relacio = '→'.join(chain)
        celorszag = next((c for c in chain if c not in ('HU', '??')), 'belföld')

        # Megbízók + megbízónkénti díj-bontás (a megbízó-elemzéshez)
        megbizo_dij = {}
        if 'Megbízó' in legs_df.columns:
            for m, s in legs_df.groupby(legs_df['Megbízó'].astype(str))['Díj részarány (EUR)']:
                megbizo_dij[m] = float(pd.to_numeric(s, errors='coerce').fillna(0).sum())
        all_megbizok = ' | '.join(megbizo_dij.keys())

        # Rögzített fuvarfeladat típusok a körben (v4.7) – időrendi lábsorrendben,
        # ismétlés nélkül; így az ügyfél a kimeneten látja, mit lát a rendszer.
        if 'Fuvarfeladat típusa' in legs_sorted_for_expl.columns:
            tipusok = ' | '.join(dict.fromkeys(
                str(t).strip() for t in legs_sorted_for_expl['Fuvarfeladat típusa']
                if str(t).strip() and str(t).strip().lower() != 'nan'))
        else:
            tipusok = ''
        # Javasolt típus-javítás a törzs-elemzésből (ha van)
        javaslatok = []
        for _t in torzsek_a_korben:
            _a = torzs_analysis_map.get(_t) or {}
            _j = str(_a.get('javaslat') or '').strip()
            if _j and _j not in javaslatok:
                javaslatok.append(_j)
        javasolt_tipus = ' | '.join(javaslatok)

        # Hónaphatár-átnyúlás (v4.7): a kör a Kör vége hónapjához van rendelve,
        # de ha a Kör kezdete korábbi hónapra esik, azt külön jelöljük.
        # v4.7: harmadik országos zárás – a kör utolsó (időrendi) lába
        # "Harmadik országba szállítás" típusú ÉS nem magyar állomáson végződik.
        harmadik_zaras = False
        if len(legs_sorted_for_expl) and 'Fuvarfeladat típusa' in legs_sorted_for_expl.columns:
            _utolso = legs_sorted_for_expl.iloc[-1]
            harmadik_zaras = (
                str(_utolso.get('Fuvarfeladat típusa') or '')
                .lower().startswith('harmadik országba')
                and not is_hu_address(_utolso.get('Utolsó Leadási állomás cím'))
            )

        kezdet_ho = (f'{pd.Timestamp(kor_kezd):%Y-%m}' if pd.notna(kor_kezd) else '')
        vege_ho = (f'{pd.Timestamp(kor_veg):%Y-%m}' if pd.notna(kor_veg) else '')
        atnyulo = bool(kezdet_ho and vege_ho and kezdet_ho != vege_ho)

        row = {
            'Kör ID': kor_id,
            'Vontatmány': vontatmany,
            'Vontatók': all_vontatok,
            'Megbízók': all_megbizok,
            'Reláció': relacio,
            'Célország': celorszag,
            'Fuvarfeladat típusok': tipusok,
            'Javasolt típus-javítás': javasolt_tipus,
            'Járatszámok': all_jaratszamok,
            'Kör kezdete dátum': kor_kezd,
            'Kör vége dátum': kor_veg,
            'Kör kezdete hónap': kezdet_ho,
            'Átnyúló kör': 'igen' if atnyulo else '',
            'Kifelé kezdő időkapu': kif_kezd_ido,
            'Kifelé kezdő cím': kif_kezd_cim,
            'Kifelé záró időkapu': kif_zar_ido,
            'Kifelé záró cím': kif_zar_cim,
            'Nemzetközi (semleges) kezdő időkapu': sem_kezd_ido,
            'Nemzetközi (semleges) kezdő cím': sem_kezd_cim,
            'Nemzetközi (semleges) záró időkapu': sem_zar_ido,
            'Nemzetközi (semleges) záró cím': sem_zar_cim,
            'Befelé kezdő időkapu': bef_kezd_ido,
            'Befelé kezdő cím': bef_kezd_cim,
            'Befelé záró időkapu': bef_zar_ido,
            'Befelé záró cím': bef_zar_cim,
            'Részfeladatok száma': len(legs_df),
            'Fuvarszámok': ', '.join(legs_df['Fuvarszám'].astype(str).tolist()),
            'Összes díj részarány (EUR)': total_dij,
            'Deviza': 'EUR',
            '_Has_kifele': has_kifele,
            '_Has_befele': has_befele,
            '_Has_semleges': has_semleges,
            '_Has_korfuvar': has_korfuvar,
            '_Implies_complete': implies_complete,
            '_Idoszak_utan_zarodo': idoszak_utan_zarodo,
            '_Idoszak_elott_kezdodo': bool(idoszak_elott_kezdodo),
            '_Ures_visszafutas_gyanu': ures_gyanu,
            '_Atnyulo_kor': atnyulo,
            '_Harmadik_orszagos_zaras': harmadik_zaras,
            '_Korben_Fuvarszam_lista': legs_df['Fuvarszám'].astype(str).tolist(),
            '_Korben_Jaratszam_lista': jaratszamok_lista,
            '_Korben_Torzs_lista': torzsek_a_korben,
            '_Korben_Megbizo_dij': megbizo_dij,
            '_Kor_Legs_Ordered': legs_ordered_tuples,
        }
        output_rows.append(row)

    result_df_all = pd.DataFrame(output_rows)

    # Magyarázat + szín (részletes: vontatmány / típus / tartalom)
    magy = []
    szin = []
    sorrend_warns = []
    for _, row in result_df_all.iterrows():
        legs_ordered = row.get('_Kor_Legs_Ordered') or []
        torzsek_a_korben = row.get('_Korben_Torzs_lista') or []
        sorrend_warn = any(str(t) in _ORDER_RESOLVED_TORZSEK for t in torzsek_a_korben)
        exp, color = build_full_kor_explanation(
            legs_ordered=legs_ordered,
            torzsek_a_korben=torzsek_a_korben,
            torzs_history_map=torzs_history_map,
            torzs_group_map=torzs_group_map,
            ures_visszafutas_gyanu=bool(row.get('_Ures_visszafutas_gyanu')),
            implies_complete=bool(row.get('_Implies_complete')),
            idoszak_utan_zarodo=bool(row.get('_Idoszak_utan_zarodo')),
            idoszak_elott_kezdodo=bool(row.get('_Idoszak_elott_kezdodo')),
            window_end=window_end,
            window_start=window_start,
            harmadik_orszagos_zaras=bool(row.get('_Harmadik_orszagos_zaras')),
        )
        magy.append(exp)
        szin.append(color)
        sorrend_warns.append('igen' if sorrend_warn else '')
    result_df_all['Sorrend-figyelmeztetés'] = sorrend_warns
    result_df_all['Magyarázat'] = magy
    result_df_all['Magyarázat_szín'] = szin
    return result_df_all


# ---------------------------------------------------------------------------
# Elemző táblák: összesítő, típushibák, ország-relációk, megbízók (v4.2)
# ---------------------------------------------------------------------------
def build_tipushiba_table(result_df: pd.DataFrame, df: pd.DataFrame) -> pd.DataFrame:
    """A kiválasztott körökben érintett törzsek típus-/állomáshibái strukturáltan
    (forrásrendszerben javítandó lista)."""
    torzs_group_map = {
        t: g for t, g in df.assign(
            _torzs=df['Fuvarszám'].astype(str).map(_torzs_of)
        ).groupby('_torzs')
    }
    _has_dt = 'Utolsó Leadási állomás időkapu (dátum)' in df.columns
    window_end = df['Utolsó Leadási állomás időkapu (dátum)'].max() if _has_dt else None
    window_start = df['Utolsó Leadási állomás időkapu (dátum)'].min() if _has_dt else None
    rows = []
    seen = set()
    for _, r in result_df.iterrows():
        for t in (r.get('_Korben_Torzs_lista') or []):
            if t in seen:
                continue
            seen.add(t)
            res = analyze_torzs_type(t, torzs_group_map.get(t), window_end=window_end,
                                     window_start=window_start)
            if res is None:
                continue
            rows.append({
                'Törzs': res['torzs'],
                'Kör ID': r['Kör ID'],
                'Kategória': ('időszak előtt kezdődő (nem hiba)'
                              if res.get('window_start_truncated')
                              else 'időszak után záródó (nem hiba)'
                              if res.get('window_truncated') else 'javítandó hiba'),
                'Fuvarszámok': res['fuvarszamok'],
                'Járatszámok': res['jaratszamok'],
                'Jelenlegi típus': res['tipus'],
                'Viszonylat': res['viszonylat'],
                'Hiba leírása': res['hiba'],
                'Javasolt javítás': res['javaslat'],
                'Kör a valóságban lezárult': 'igen' if res['implies_complete'] else 'nem',
            })
    return pd.DataFrame(rows)


def build_osszesito_table(result_df: pd.DataFrame) -> pd.DataFrame:
    """Összesítő tábla: kör-darabszámok kategóriánként + bevétel/költség/eredmény szummák.

    v4.6: az 'időszak után záródó' (kék) körök a pénzügyi mutatókból KIZÁRVA."""
    szin_all = result_df.get('Magyarázat_szín', pd.Series('', index=result_df.index)).astype(str)
    kizart_mask = szin_all.str.contains('lightblue')
    base = result_df[~kizart_mask]
    szin = base.get('Magyarázat_szín', pd.Series('', index=base.index)).astype(str)

    n = len(result_df)
    n_kizart = int(kizart_mask.sum())
    n_zold = int(szin.str.contains('lightgreen').sum())
    n_piros = int(szin.str.contains('lightcoral').sum())
    n_narancs = len(base) - n_zold - n_piros

    bev = pd.to_numeric(base.get('Összes díj részarány (EUR)'), errors='coerce')
    tk = pd.to_numeric(base.get('Teljes költség'), errors='coerce') \
        if 'Teljes költség' in base.columns else pd.Series(dtype=float)
    er = pd.to_numeric(base.get('Járati eredmény'), errors='coerce') \
        if 'Járati eredmény' in base.columns else pd.Series(dtype=float)

    rows = [
        ('Generálta', f'Bábolna Körfuvar Generálás {APP_VERSION} ({APP_RELEASE_DATE})'),
        ('Körök összesen (db)', n),
        ('Teljes kör – zöld (db)', n_zold),
        ('Részleges / hibás – narancs (db)', n_narancs),
        ('Szétesett törzs – piros (db)', n_piros),
        ('Időszak határán csonkolt – kék, pénzügyi kimutatásból kizárva (db)', n_kizart),
        ('  ebből időszak UTÁN záródó (db)',
         int(result_df.get('_Idoszak_utan_zarodo', pd.Series(dtype=bool)).fillna(False).astype(bool).sum())
         if '_Idoszak_utan_zarodo' in result_df.columns else 0),
        ('  ebből időszak ELŐTT kezdődő (db)',
         int(result_df.get('_Idoszak_elott_kezdodo', pd.Series(dtype=bool)).fillna(False).astype(bool).sum())
         if '_Idoszak_elott_kezdodo' in result_df.columns else 0),
        ('Összes bevétel – díj részarány (EUR)', round(float(bev.sum()), 2) if len(bev) else 0),
        ('Összes költség (EUR)', round(float(tk.sum()), 2) if len(tk) else None),
        ('Összes járati eredmény (EUR)', round(float(er.sum()), 2) if len(er) else None),
        ('Átlagos eredmény / kör (EUR)', round(float(er.mean()), 2) if len(er) and er.notna().any() else None),
        ('Veszteséges körök (db)', int((er < 0).sum()) if len(er) else None),
        ('Nyereséges körök (db)', int((er > 0).sum()) if len(er) else None),
    ]

    # v4.7: hónaphatáron átnyúló körök (a kör az előző hónapban indult, de ebben
    # a hónapban zárult → a teljes díja ehhez a hónaphoz kerül). A besorolás nem
    # változik, csak láthatóvá tesszük, mekkora rész "csúszott át" a hónapok között.
    if 'Átnyúló kör' in base.columns:
        atny = base['Átnyúló kör'].astype(str).eq('igen')
        n_atny = int(atny.sum())
        bev_atny = pd.to_numeric(
            base.loc[atny, 'Összes díj részarány (EUR)'], errors='coerce').sum()
        bev_ossz = float(bev.sum()) if len(bev) else 0.0
        rows.extend([
            ('Előző hónapban indult, itt zárult kör (db)', n_atny),
            ('Ebből származó bevétel (EUR)', round(float(bev_atny), 2)),
            ('Ennek aránya a havi bevételből (%)',
             round(100.0 * float(bev_atny) / bev_ossz, 1) if bev_ossz else 0.0),
        ])

    return pd.DataFrame(rows, columns=['Mutató', 'Érték'])


def build_relacio_table(result_df: pd.DataFrame) -> pd.DataFrame:
    """Célország (első nem-HU ország a lánban) szerinti profit-elemzés,
    eredmény szerint csökkenő sorrendben."""
    if 'Célország' not in result_df.columns or result_df.empty:
        return pd.DataFrame()
    t = result_df.copy()
    # v4.6: időszak után záródó körök kizárva a profit-elemzésből
    if '_Idoszak_utan_zarodo' in t.columns:
        _kiz = t['_Idoszak_utan_zarodo'].fillna(False).astype(bool)
        if '_Idoszak_elott_kezdodo' in t.columns:
            _kiz = _kiz | t['_Idoszak_elott_kezdodo'].fillna(False).astype(bool)
        t = t[~_kiz]
    elif 'Magyarázat_szín' in t.columns:
        t = t[~t['Magyarázat_szín'].astype(str).str.contains('lightblue')]
    if t.empty:
        return pd.DataFrame()
    t['_bev'] = pd.to_numeric(t.get('Összes díj részarány (EUR)'), errors='coerce')
    t['_tk'] = pd.to_numeric(t.get('Teljes költség'), errors='coerce')
    t['_er'] = pd.to_numeric(t.get('Járati eredmény'), errors='coerce')
    g = t.groupby('Célország').agg(
        Körök_db=('Kör ID', 'count'),
        Bevétel_EUR=('_bev', 'sum'),
        Költség_EUR=('_tk', 'sum'),
        Eredmény_EUR=('_er', 'sum'),
        Átlag_eredmény_per_kör=('_er', 'mean'),
    ).reset_index()
    g = g.rename(columns={
        'Körök_db': 'Körök (db)', 'Bevétel_EUR': 'Bevétel (EUR)',
        'Költség_EUR': 'Költség (EUR)', 'Eredmény_EUR': 'Eredmény (EUR)',
        'Átlag_eredmény_per_kör': 'Átlag eredmény / kör (EUR)',
    })
    for c in ['Bevétel (EUR)', 'Költség (EUR)', 'Eredmény (EUR)', 'Átlag eredmény / kör (EUR)']:
        g[c] = g[c].round(2)
    return g.sort_values('Eredmény (EUR)', ascending=False).reset_index(drop=True)


def build_megbizo_table(result_df: pd.DataFrame) -> pd.DataFrame:
    """Megbízó szerinti profit-elemzés. Ha egy körben több megbízó van, a kör
    költségét/eredményét a díj részarányuk arányában osztjuk fel közöttük."""
    rows = []
    for _, r in result_df.iterrows():
        # v4.6: időszak után záródó körök kizárva a profit-elemzésből
        if bool(r.get('_Idoszak_utan_zarodo')) or bool(r.get('_Idoszak_elott_kezdodo')):
            continue
        if 'lightblue' in str(r.get('Magyarázat_szín', '')):
            continue
        md = r.get('_Korben_Megbizo_dij') or {}
        if not md:
            continue
        total_dij = sum(md.values())
        tk = pd.to_numeric(pd.Series([r.get('Teljes költség')]), errors='coerce').iloc[0]
        for m, dij in md.items():
            share = (dij / total_dij) if total_dij > 0 else 1.0 / len(md)
            alloc_tk = float(tk) * share if pd.notna(tk) else None
            rows.append({
                'Megbízó': m,
                'Kör ID': r['Kör ID'],
                'Bevétel (EUR)': dij,
                'Költség (EUR)': alloc_tk,
                'Eredmény (EUR)': (dij - alloc_tk) if alloc_tk is not None else None,
            })
    if not rows:
        return pd.DataFrame()
    t = pd.DataFrame(rows)
    g = t.groupby('Megbízó').agg(
        Körök_db=('Kör ID', 'nunique'),
        Bevétel=('Bevétel (EUR)', 'sum'),
        Költség=('Költség (EUR)', 'sum'),
        Eredmény=('Eredmény (EUR)', 'sum'),
    ).reset_index()
    g['Átlag eredmény / kör (EUR)'] = g['Eredmény'] / g['Körök_db']
    g = g.rename(columns={
        'Körök_db': 'Körök (db)', 'Bevétel': 'Bevétel (EUR)',
        'Költség': 'Költség (EUR)', 'Eredmény': 'Eredmény (EUR)',
    })
    for c in ['Bevétel (EUR)', 'Költség (EUR)', 'Eredmény (EUR)', 'Átlag eredmény / kör (EUR)']:
        g[c] = pd.to_numeric(g[c], errors='coerce').round(2)
    return g.sort_values('Eredmény (EUR)', ascending=False).reset_index(drop=True)


# ---------------------------------------------------------------------------
# Költségtábla (controlling - kategóriákkal) beolvasó – két-soros, merge-elt fejléccel
# ---------------------------------------------------------------------------
def parse_cost_file(uploaded_file):
    """Járatszám alapú költségtábla beolvasása.

    A fájl fejléce két soros:
      - 1. sor: összevont kategória-nevek (merge-elt cellák, ffill-lel expanded).
      - 2. sor: 'Költség', 'Km', 'EUR', 'Km' jelölések – csak a 'Költség' oszlopokat
                indexeljük kategóriánként; a 'Km'-ből az elsőt tartjuk meg.
    """
    raw = pd.read_excel(uploaded_file, header=None, sheet_name=0)
    if len(raw) < 3:
        return pd.DataFrame(columns=['Járatszám', 'Km']), []

    cost_row = raw.iloc[0].ffill()
    type_row = raw.iloc[1]

    cost_col_map = {}
    km_col_idx = None
    for idx in range(1, len(type_row)):
        t = type_row.iloc[idx]
        cat = cost_row.iloc[idx] if idx < len(cost_row) else None
        if t == 'Költség' and pd.notna(cat):
            cost_col_map[str(cat).strip()] = idx
        elif t == 'Km' and km_col_idx is None:
            km_col_idx = idx

    data = raw.iloc[2:].reset_index(drop=True)
    out = pd.DataFrame()
    out['Járatszám'] = data.iloc[:, 0].astype(str).str.strip()
    if km_col_idx is not None:
        out['Km'] = pd.to_numeric(data.iloc[:, km_col_idx], errors='coerce')
    else:
        out['Km'] = pd.NA

    categories = list(cost_col_map.keys())
    for cat in categories:
        out[cat] = pd.to_numeric(data.iloc[:, cost_col_map[cat]], errors='coerce')

    out = out[out['Járatszám'].str.len() > 0]
    out = out[~out['Járatszám'].isin(['nan', 'NaN', 'None'])].reset_index(drop=True)
    return out, categories


def load_all_cost_files(uploaded_files):
    """Több kategóriás költség fájl beolvasása és összefűzése.

    Ha ugyanaz a járatszám több fájlban is szerepel, csak az ELSŐ előfordulást tartjuk meg
    (keep='first'), így egy járat csak egyszer számolódik az aggregálásnál.
    """
    if not uploaded_files:
        return None, []
    dfs = []
    all_cats = []
    for f in uploaded_files:
        try:
            cf, cats = parse_cost_file(f)
        except Exception as e:
            st.warning(f"A(z) '{getattr(f, 'name', '?')}' költség fájl beolvasása sikertelen: {e}")
            continue
        if cf is None or cf.empty:
            continue
        dfs.append(cf)
        for c in cats:
            if c not in all_cats:
                all_cats.append(c)
    if not dfs:
        return None, []
    combined = pd.concat(dfs, ignore_index=True, sort=False)
    combined = combined.drop_duplicates(subset='Járatszám', keep='first').reset_index(drop=True)
    return combined, all_cats


# ---------------------------------------------------------------------------
# Járat-alapú eredménykimutatás (flight controlling) beolvasó – egysoros fejléc
# ---------------------------------------------------------------------------
FC_EXTRA_COST_COLS = ['Gázolaj költség', 'Adblue költség', 'Útdíj költség']
FC_RUN_COLS = [
    'Menetlevél ∑ nap',
    'Menetlevél ∑ óra',
    'Menetlevél ∑ km',
    'Rakott km',
    'Üres km',
    'Menetlevél tankolás',
]
FC_DERIVED_COLS = ['Rakott km %', 'Üres km %', 'Menetlevél tankolás / 100 km']


def parse_flight_controlling_file(uploaded_file):
    """Egy flight controlling riport beolvasása, csak a szükséges oszlopokat megtartva."""
    raw = pd.read_excel(uploaded_file, sheet_name=0)
    if raw.empty or 'Járatszám' not in raw.columns:
        return pd.DataFrame()

    wanted = ['Járatszám'] + FC_EXTRA_COST_COLS + FC_RUN_COLS
    cols = [c for c in wanted if c in raw.columns]
    out = raw[cols].copy()
    out['Járatszám'] = out['Járatszám'].astype(str).str.strip()
    for c in cols[1:]:
        out[c] = pd.to_numeric(out[c], errors='coerce')
    out = out[out['Járatszám'].str.len() > 0]
    out = out[~out['Járatszám'].isin(['nan', 'NaN', 'None'])].reset_index(drop=True)
    return out


def load_all_flight_controlling_files(uploaded_files):
    """Több flight controlling fájl beolvasása + dedup Járatszám alapján (keep='first')."""
    if not uploaded_files:
        return None
    dfs = []
    for f in uploaded_files:
        try:
            fc = parse_flight_controlling_file(f)
        except Exception as e:
            st.warning(f"A(z) '{getattr(f, 'name', '?')}' eredménykimutatás fájl beolvasása sikertelen: {e}")
            continue
        if fc is None or fc.empty:
            continue
        dfs.append(fc)
    if not dfs:
        return None
    combined = pd.concat(dfs, ignore_index=True, sort=False)
    combined = combined.drop_duplicates(subset='Járatszám', keep='first').reset_index(drop=True)
    return combined


# ---------------------------------------------------------------------------
# Aggregáció körökre
# ---------------------------------------------------------------------------
def aggregate_cost_for_rings(result_df: pd.DataFrame,
                             cost_df: pd.DataFrame,
                             cost_categories: list):
    """A körök járatszámai alapján aggregálja a (kategóriás) költségtáblát körönként.

    Hozzáadott oszlopok:
      - 'Km'
      - minden 'cost_categories' elem

    A 'Teljes költség' és 'Járati eredmény' számítását NEM itt végezzük, azt a
    flight-controlling aggregáció után tesszük egyszerre, hogy az összes költség
    összegződjön.
    """
    cost_map = cost_df.set_index('Járatszám').to_dict(orient='index')

    km_list = []
    cat_cols = {c: [] for c in cost_categories}
    any_match_list = []

    for _, row in result_df.iterrows():
        kor_jaratok = row.get('_Korben_Jaratszam_lista') or []
        seen = set()
        km_sum = 0.0
        cat_sums = {c: 0.0 for c in cost_categories}
        any_match = False

        for j in kor_jaratok:
            jc = str(j).strip()
            if jc in seen:
                continue
            seen.add(jc)
            entry = cost_map.get(jc)
            if entry is None:
                continue
            any_match = True
            kv = entry.get('Km')
            if pd.notna(kv):
                km_sum += float(kv)
            for c in cost_categories:
                v = entry.get(c)
                if pd.notna(v):
                    cat_sums[c] += float(v)

        if any_match:
            km_list.append(km_sum)
            for c in cost_categories:
                cat_cols[c].append(cat_sums[c])
        else:
            km_list.append(None)
            for c in cost_categories:
                cat_cols[c].append(None)
        any_match_list.append(any_match)

    result_df = result_df.copy()
    result_df['Km'] = km_list
    for c in cost_categories:
        result_df[c] = cat_cols[c]
    result_df['_cost_any_match'] = any_match_list
    return result_df


def aggregate_flight_controlling_for_rings(result_df: pd.DataFrame,
                                           fc_df: pd.DataFrame):
    """Flight controlling riport (járat alapú) aggregálása körökre.

    Hozzáadja a körhöz:
      - Gázolaj költség, Adblue költség, Útdíj költség (szumma)
      - Menetlevél ∑ nap, ∑ óra, ∑ km, Rakott km, Üres km, Menetlevél tankolás (szumma)
      - Rakott km %, Üres km %, Menetlevél tankolás / 100 km (számított)
    """
    available_cost_cols = [c for c in FC_EXTRA_COST_COLS if c in fc_df.columns]
    available_run_cols = [c for c in FC_RUN_COLS if c in fc_df.columns]
    tracked = available_cost_cols + available_run_cols

    fc_map = fc_df.set_index('Járatszám').to_dict(orient='index')
    out_cols = {c: [] for c in tracked}
    any_match_list = []

    for _, row in result_df.iterrows():
        kor_jaratok = row.get('_Korben_Jaratszam_lista') or []
        seen = set()
        sums = {c: 0.0 for c in tracked}
        any_match = False
        for j in kor_jaratok:
            jc = str(j).strip()
            if jc in seen:
                continue
            seen.add(jc)
            entry = fc_map.get(jc)
            if entry is None:
                continue
            any_match = True
            for c in tracked:
                v = entry.get(c)
                if pd.notna(v):
                    sums[c] += float(v)
        if any_match:
            for c in tracked:
                out_cols[c].append(sums[c])
        else:
            for c in tracked:
                out_cols[c].append(None)
        any_match_list.append(any_match)

    result_df = result_df.copy()
    for c in tracked:
        result_df[c] = out_cols[c]
    result_df['_fc_any_match'] = any_match_list

    # Számított oszlopok
    def _safe_div_pct(num, denom):
        if pd.isna(num) or pd.isna(denom) or denom in (0, 0.0):
            return None
        try:
            return float(num) / float(denom) * 100.0
        except (ValueError, TypeError):
            return None

    if 'Rakott km' in result_df.columns and 'Menetlevél ∑ km' in result_df.columns:
        result_df['Rakott km %'] = [
            _safe_div_pct(r, k) for r, k in zip(result_df['Rakott km'], result_df['Menetlevél ∑ km'])
        ]
    if 'Üres km' in result_df.columns and 'Menetlevél ∑ km' in result_df.columns:
        result_df['Üres km %'] = [
            _safe_div_pct(u, k) for u, k in zip(result_df['Üres km'], result_df['Menetlevél ∑ km'])
        ]
    # Menetlevél tankolás / 100 km = Menetlevél tankolás / (Menetlevél ∑ km / 100)
    #   (azaz L/100 km: a megtett 100 km-re jutó tankolás mennyisége)
    if 'Menetlevél tankolás' in result_df.columns and 'Menetlevél ∑ km' in result_df.columns:
        vals = []
        for t, k in zip(result_df['Menetlevél tankolás'], result_df['Menetlevél ∑ km']):
            if pd.isna(t) or pd.isna(k):
                vals.append(None)
                continue
            try:
                k_f = float(k)
                if k_f == 0:
                    vals.append(None)
                else:
                    vals.append(float(t) / (k_f / 100.0))
            except (ValueError, TypeError, ZeroDivisionError):
                vals.append(None)
        result_df['Menetlevél tankolás / 100 km'] = vals

    return result_df


def finalize_totals(result_df: pd.DataFrame, all_cost_cols: list):
    """Teljes költség = az összegzendő költség oszlopok sorbeli szumma.
    Járati eredmény = Összes díj részarány - Teljes költség.

    FONTOS: ha az FC 'Útdíj költség' oszlop jelen van a summában, akkor az első
    (kategóriás) fájl ország-alapú kategóriái (Magyar/Belga/Francia/... = az útdíj
    ország-bontása) NEM adódnak hozzá a szummához – különben duplikálódna az útdíj.
    A kategóriás oszlopok maguk MEGMARADNAK a táblázatban, csak a Teljes költség
    számításából vannak kizárva ilyenkor.

    Ha sem a kategóriás költségtábla, sem a flight-controlling nem adott költséget,
    akkor mindkét mező marad NaN – nem tévesztjük meg zöld színnel a sor kezelését.
    """
    result_df = result_df.copy()
    has_any_cost_source = ('_cost_any_match' in result_df.columns) or ('_fc_any_match' in result_df.columns)

    if not all_cost_cols:
        result_df['Teljes költség'] = None
    else:
        # Duplikáció-védelem: ha az FC-Útdíj költség jelen van, az ország-bontás nem
        # kerül a szummába (mert az FC-Útdíj = ország-bontás szummája).
        has_fc_utdij = ('Útdíj költség' in all_cost_cols) and ('Útdíj költség' in result_df.columns)
        if has_fc_utdij:
            summable_cols = [c for c in all_cost_cols if c not in COUNTRY_TOLL_CATEGORIES]
        else:
            summable_cols = list(all_cost_cols)

        if summable_cols:
            num = result_df[summable_cols].apply(pd.to_numeric, errors='coerce')
            sum_val = num.sum(axis=1, min_count=1)
        else:
            sum_val = pd.Series([None] * len(result_df))

        if has_any_cost_source:
            any_cost_any = pd.Series([False] * len(result_df))
            if '_cost_any_match' in result_df.columns:
                any_cost_any = any_cost_any | result_df['_cost_any_match'].fillna(False).astype(bool)
            if '_fc_any_match' in result_df.columns:
                any_cost_any = any_cost_any | result_df['_fc_any_match'].fillna(False).astype(bool)
            sum_val = sum_val.where(any_cost_any)
        result_df['Teljes költség'] = sum_val

    dij = pd.to_numeric(result_df.get('Összes díj részarány (EUR)'), errors='coerce')
    tk = pd.to_numeric(result_df['Teljes költség'], errors='coerce')
    result_df['Járati eredmény'] = dij - tk
    return result_df


# ---------------------------------------------------------------------------
# CSS → openpyxl PatternFill
# ---------------------------------------------------------------------------
def _css_to_openpyxl_fill(css: str):
    """CSS háttér stringből openpyxl PatternFill. None, ha nem ismert."""
    if not isinstance(css, str):
        return None
    if 'lightgreen' in css:
        return PatternFill('solid', start_color='C6EFCE')
    if 'lightblue' in css:
        return PatternFill('solid', start_color='BDD7EE')
    if 'orange' in css:
        return PatternFill('solid', start_color='FFD699')
    if 'lightcoral' in css:
        return PatternFill('solid', start_color='FFC7CE')
    return None


# ---------------------------------------------------------------------------
# Csoport színek a fejléchez (futási / költség / bevétel / eredmény)
# ---------------------------------------------------------------------------
GROUP_COLORS = {
    'futas':    {'hex': 'D9E1F2', 'css': 'background-color: #D9E1F2; color: black'},  # halványkék
    'koltseg':  {'hex': 'FCE4D6', 'css': 'background-color: #FCE4D6; color: black'},  # halvány barack
    'bevetel':  {'hex': 'E2EFDA', 'css': 'background-color: #E2EFDA; color: black'},  # halványzöld
    'eredmeny': {'hex': 'FFE699', 'css': 'background-color: #FFE699; color: black'},  # halványsárga
}

# Feldolgozási animáció: guruló kamion (tiszta CSS, extra függőség nélkül)
TRUCK_ANIM_HTML = """
<div class="truck-scene">
  <div class="truck">🚛</div>
  <div class="cargo">📦</div>
  <div class="road"></div>
</div>
<style>
.truck-scene{position:relative;height:78px;overflow:hidden;margin:4px 0 10px 0;}
.truck{position:absolute;bottom:14px;font-size:44px;line-height:1;
       animation:drive 3.2s linear infinite;transform:scaleX(-1);}
.cargo{position:absolute;bottom:22px;font-size:20px;opacity:0;
       animation:cargo 3.2s linear infinite;}
.road{position:absolute;bottom:8px;height:5px;width:100%;border-radius:3px;
      background:repeating-linear-gradient(90deg,#9aa0a6 0 32px,transparent 32px 54px);
      animation:roadmove .5s linear infinite;}
@keyframes drive{0%{left:-14%}100%{left:104%}}
@keyframes cargo{0%,55%{opacity:0;left:60%}60%{opacity:1;left:60%;bottom:22px}
                 100%{opacity:0;left:60%;bottom:44px}}
@keyframes roadmove{0%{background-position:0 0}100%{background-position:-54px 0}}
</style>
"""

IDENTITY_COLS = [
    'Kör ID', 'Vontatmány', 'Vontatók', 'Megbízók',
    'Reláció', 'Célország', 'Fuvarfeladat típusok', 'Javasolt típus-javítás',
    'Járatszámok',
    'Kör kezdete dátum', 'Kör vége dátum', 'Kör kezdete hónap', 'Átnyúló kör',
    'Kifelé kezdő időkapu', 'Kifelé kezdő cím',
    'Kifelé záró időkapu', 'Kifelé záró cím',
    'Nemzetközi (semleges) kezdő időkapu', 'Nemzetközi (semleges) kezdő cím',
    'Nemzetközi (semleges) záró időkapu', 'Nemzetközi (semleges) záró cím',
    'Befelé kezdő időkapu', 'Befelé kezdő cím',
    'Befelé záró időkapu', 'Befelé záró cím',
    'Részfeladatok száma', 'Fuvarszámok',
    'Sorrend-figyelmeztetés',
    'Magyarázat',
]


def order_and_group_columns(result_df: pd.DataFrame, cost_categories: list):
    """A megjelenítési oszlopsorrend és az oszlop→csoport map összeállítása."""
    futas_cols = []
    if 'Km' in result_df.columns:
        futas_cols.append('Km')
    for c in ['Menetlevél ∑ nap', 'Menetlevél ∑ óra', 'Menetlevél ∑ km',
              'Rakott km', 'Rakott km %', 'Üres km', 'Üres km %',
              'Menetlevél tankolás', 'Menetlevél tankolás / 100 km']:
        if c in result_df.columns:
            futas_cols.append(c)

    koltseg_cols = []
    for c in cost_categories:
        if c in result_df.columns:
            koltseg_cols.append(c)
    for c in FC_EXTRA_COST_COLS:
        if c in result_df.columns and c not in koltseg_cols:
            koltseg_cols.append(c)
    if 'Teljes költség' in result_df.columns:
        koltseg_cols.append('Teljes költség')

    bevetel_cols = [c for c in ['Összes díj részarány (EUR)', 'Deviza'] if c in result_df.columns]
    eredmeny_cols = [c for c in ['Járati eredmény'] if c in result_df.columns]

    ordered_cols = [c for c in IDENTITY_COLS if c in result_df.columns]
    ordered_cols.extend(futas_cols)
    ordered_cols.extend(koltseg_cols)
    ordered_cols.extend(bevetel_cols)
    ordered_cols.extend(eredmeny_cols)

    group_of_col = {}
    for c in futas_cols:
        group_of_col[c] = 'futas'
    for c in koltseg_cols:
        group_of_col[c] = 'koltseg'
    for c in bevetel_cols:
        group_of_col[c] = 'bevetel'
    for c in eredmeny_cols:
        group_of_col[c] = 'eredmeny'

    return ordered_cols, group_of_col, futas_cols, koltseg_cols


def build_month_xlsx(result_df_display: pd.DataFrame, group_of_col: dict,
                     sum_cols: set, extra_sheets: list) -> bytes:
    """Egy havi körfuvar-Excel összeállítása (fő munkalap + szumma sor + segédfülek).
    extra_sheets: [(df, munkalap_név), ...]"""
    xlsx_df = result_df_display.drop(columns=['Magyarázat_szín'], errors='ignore')
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        xlsx_df.to_excel(writer, index=False, sheet_name='körfuvarok')
        ws = writer.sheets['körfuvarok']

        header_to_col = {}
        for c in range(1, ws.max_column + 1):
            header_to_col[ws.cell(row=1, column=c).value] = c

        # Fejléc sor csoport-szín
        for col_name, col_idx in header_to_col.items():
            g = group_of_col.get(col_name)
            if g:
                ws.cell(row=1, column=col_idx).fill = PatternFill(
                    'solid', start_color=GROUP_COLORS[g]['hex'])

        # Járati eredmény: sor-szintű zöld/piros
        er_col = header_to_col.get('Járati eredmény')
        if er_col is not None:
            green_fill = PatternFill('solid', start_color='C6EFCE')
            red_fill = PatternFill('solid', start_color='FFC7CE')
            for r in range(2, ws.max_row + 1):
                v = ws.cell(row=r, column=er_col).value
                if isinstance(v, (int, float)):
                    if v < 0:
                        ws.cell(row=r, column=er_col).fill = red_fill
                    elif v > 0:
                        ws.cell(row=r, column=er_col).fill = green_fill

        # Magyarázat: sor-szintű szín
        magy_col = header_to_col.get('Magyarázat')
        if magy_col is not None and 'Magyarázat_szín' in result_df_display.columns:
            for i, css in enumerate(result_df_display['Magyarázat_szín'].tolist()):
                fill = _css_to_openpyxl_fill(css)
                if fill is not None:
                    ws.cell(row=i + 2, column=magy_col).fill = fill

        # Szumma sor (bevétel / költségek / eredmény / futási adatok)
        # v4.6: az 'időszak után záródó' (kék) körök a szummából kizárva
        if 'Magyarázat_szín' in result_df_display.columns:
            _kizart = result_df_display['Magyarázat_szín'].astype(str) \
                .str.contains('lightblue').values
        else:
            _kizart = pd.Series(False, index=xlsx_df.index).values
        _sum_base = xlsx_df[~_kizart]
        sum_fill = PatternFill('solid', start_color='DDDDDD')
        sum_row_idx = ws.max_row + 1
        ws.cell(row=sum_row_idx, column=1).value = (
            'ÖSSZESEN (időszak után záródó körök nélkül)' if _kizart.any() else 'ÖSSZESEN'
        )
        ws.cell(row=sum_row_idx, column=1).fill = sum_fill
        for col_name, col_idx in header_to_col.items():
            if col_name in sum_cols and col_name in _sum_base.columns:
                s = pd.to_numeric(_sum_base[col_name], errors='coerce').sum()
                if pd.notna(s):
                    c = ws.cell(row=sum_row_idx, column=col_idx)
                    c.value = round(float(s), 2)
                    c.fill = sum_fill

        # Segéd munkalapok
        head_fill = PatternFill('solid', start_color='D9E1F2')
        for df_sheet, name in extra_sheets:
            if df_sheet is None or len(df_sheet) == 0:
                continue
            df_sheet.to_excel(writer, index=False, sheet_name=name)
            ws2 = writer.sheets[name]
            for ci in range(1, ws2.max_column + 1):
                ws2.cell(row=1, column=ci).fill = head_fill

    buffer.seek(0)
    return buffer.getvalue()


def build_havi_osszesito_table(month_results: dict) -> pd.DataFrame:
    """Hónaponkénti összesítő tábla (egy sor = egy hónap) + ÖSSZESEN sor.
    month_results: {label: finalizált result_df}"""
    rows = []
    for label, rdf in month_results.items():
        szin_all = rdf.get('Magyarázat_szín', pd.Series('', index=rdf.index)).astype(str)
        kizart_mask = szin_all.str.contains('lightblue')
        base = rdf[~kizart_mask]
        szin = base.get('Magyarázat_szín', pd.Series('', index=base.index)).astype(str)
        bev = pd.to_numeric(base.get('Összes díj részarány (EUR)'), errors='coerce')
        tk = pd.to_numeric(base.get('Teljes költség'), errors='coerce') \
            if 'Teljes költség' in base.columns else pd.Series(dtype=float)
        er = pd.to_numeric(base.get('Járati eredmény'), errors='coerce') \
            if 'Járati eredmény' in base.columns else pd.Series(dtype=float)
        rows.append({
            'Hónap': label,
            'Körök (db)': len(rdf),
            'Teljes kör (db)': int(szin.str.contains('lightgreen').sum()),
            'Részleges/hibás (db)': int((~szin.str.contains('lightgreen')).sum()),
            'Időszak határán csonkolt – kizárva (db)': int(kizart_mask.sum()),
            'Előző hónapban indult kör (db)': int(
                base['Átnyúló kör'].astype(str).eq('igen').sum())
            if 'Átnyúló kör' in base.columns else 0,
            'Előző hónapból áthozott bevétel (EUR)': round(float(pd.to_numeric(
                base.loc[base['Átnyúló kör'].astype(str).eq('igen'),
                         'Összes díj részarány (EUR)'], errors='coerce').sum()), 2)
            if 'Átnyúló kör' in base.columns else 0,
            'Bevétel (EUR)': round(float(bev.sum()), 2) if len(bev) else 0,
            'Költség (EUR)': round(float(tk.sum()), 2) if len(tk) and tk.notna().any() else None,
            'Eredmény (EUR)': round(float(er.sum()), 2) if len(er) and er.notna().any() else None,
            'Átlag eredmény / kör (EUR)': round(float(er.mean()), 2) if len(er) and er.notna().any() else None,
        })
    out = pd.DataFrame(rows)
    if len(out) > 1:
        total = {
            'Hónap': 'ÖSSZESEN',
            'Körök (db)': int(out['Körök (db)'].sum()),
            'Teljes kör (db)': int(out['Teljes kör (db)'].sum()),
            'Részleges/hibás (db)': int(out['Részleges/hibás (db)'].sum()),
            'Időszak határán csonkolt – kizárva (db)': int(out['Időszak határán csonkolt – kizárva (db)'].sum()),
            'Előző hónapban indult kör (db)': int(out['Előző hónapban indult kör (db)'].sum()),
            'Előző hónapból áthozott bevétel (EUR)': round(float(pd.to_numeric(
                out['Előző hónapból áthozott bevétel (EUR)'], errors='coerce').sum()), 2),
            'Bevétel (EUR)': round(float(pd.to_numeric(out['Bevétel (EUR)'], errors='coerce').sum()), 2),
            'Költség (EUR)': round(float(pd.to_numeric(out['Költség (EUR)'], errors='coerce').sum()), 2),
            'Eredmény (EUR)': round(float(pd.to_numeric(out['Eredmény (EUR)'], errors='coerce').sum()), 2),
            'Átlag eredmény / kör (EUR)': None,
        }
        out = pd.concat([out, pd.DataFrame([total])], ignore_index=True)
    return out


def build_summary_xlsx(month_results: dict, tipushiba_all: pd.DataFrame,
                       datum_anomalia_df: pd.DataFrame) -> bytes:
    """Összesítő munkafüzet a legyártott hónapokról: havi összehasonlítás +
    a teljes időszakra aggregált ország-reláció és megbízó elemzés."""
    havi_df = build_havi_osszesito_table(month_results)
    combined = pd.concat(month_results.values(), ignore_index=True)
    # Körfuvar-only módban (nincs költség-adat) a profit-elemzés fülek kimaradnak
    if 'Teljes költség' in combined.columns:
        relacio_df = build_relacio_table(combined)
        megbizo_df = build_megbizo_table(combined)
    else:
        relacio_df = pd.DataFrame()
        megbizo_df = pd.DataFrame()

    buffer = io.BytesIO()
    head_fill = PatternFill('solid', start_color='D9E1F2')
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        for df_sheet, name in [
            (havi_df, 'Havi összesítő'),
            (relacio_df, 'Ország-relációk (időszak)'),
            (megbizo_df, 'Megbízók (időszak)'),
            (tipushiba_all, 'Típushibák (javítandó)'),
            (datum_anomalia_df, 'Dátum-anomáliák'),
        ]:
            if df_sheet is None or len(df_sheet) == 0:
                continue
            df_sheet.to_excel(writer, index=False, sheet_name=name)
            ws = writer.sheets[name]
            for ci in range(1, ws.max_column + 1):
                ws.cell(row=1, column=ci).fill = head_fill
            if name == 'Havi összesítő':
                sum_fill = PatternFill('solid', start_color='DDDDDD')
                for ci in range(1, ws.max_column + 1):
                    if str(ws.cell(row=ws.max_row, column=1).value) == 'ÖSSZESEN':
                        ws.cell(row=ws.max_row, column=ci).fill = sum_fill
    buffer.seek(0)
    return buffer.getvalue()


# ---------------------------------------------------------------------------
# Streamlit UI
# ---------------------------------------------------------------------------
uploaded_logbooks = st.file_uploader(
    'Válaszd ki a fuvarnapló Excel fájl(oka)t – több időszak is feltölthető egyszerre',
    type=['xlsx'],
    accept_multiple_files=True,
)

cost_files = st.file_uploader(
    'Válaszd ki a költség Excel fájl(oka)t – opcionális, több is lehet',
    type=['xlsx'],
    accept_multiple_files=True,
)

fc_files = st.file_uploader(
    'Válaszd ki a járat-alapú eredménykimutatás fájl(oka)t – opcionális, több is lehet',
    type=['xlsx'],
    accept_multiple_files=True,
)


if uploaded_logbooks:
    _dfs = []
    for _f in uploaded_logbooks:
        try:
            _dfs.append(pd.read_excel(_f))
        except Exception as _e:
            st.warning(f"A(z) '{getattr(_f, 'name', '?')}' fuvarnapló beolvasása sikertelen: {_e}")
    if not _dfs:
        st.error('Egyik fuvarnapló fájl sem olvasható.')
        st.stop()

    df = pd.concat(_dfs, ignore_index=True, sort=False)
    _n_before = len(df)
    # Átfedő időszakú fájloknál ugyanaz a részfeladat többször szerepelhet –
    # Fuvarszám alapján deduplikálunk (az első előfordulás marad).
    _elteres_fuvarok = []
    if 'Fuvarszám' in df.columns:
        # Eltérő tartalmú duplikátumok felderítése: ha ugyanaz a Fuvarszám a két
        # fájlban NEM azonos tartalommal szerepel (pl. módosították a forrásban
        # a két export között), azt jelezzük, mert a dedup a feltöltési sorrend
        # szerinti ELSŐ verziót tartja meg.
        _dup_mask = df.duplicated(subset='Fuvarszám', keep=False)
        if _dup_mask.any():
            _dups = df[_dup_mask].astype(str)
            for _fsz, _grp in _dups.groupby('Fuvarszám'):
                if len(_grp.drop_duplicates()) > 1:
                    _elteres_fuvarok.append(_fsz)
        df = df.drop_duplicates(subset='Fuvarszám', keep='first').reset_index(drop=True)
    _n_dup = _n_before - len(df)
    st.success(
        f'✅ Fuvarnapló betöltve: {len(uploaded_logbooks)} fájl, {len(df)} egyedi sor'
        + (f' ({_n_dup} átfedő sor kiszűrve)' if _n_dup else '')
    )
    if _elteres_fuvarok:
        st.warning(
            f'⚠️ {len(_elteres_fuvarok)} fuvar ELTÉRŐ tartalommal szerepel a feltöltött '
            f'fájlokban (a forrásban módosíthatták a két export között). A feldolgozás a '
            f'feltöltési sorrend szerinti első verziót használja — érdemes a legfrissebb '
            f'exportot feltölteni elsőként, vagy ellenőrizni ezeket: '
            + ', '.join(_elteres_fuvarok[:20])
            + (' …' if len(_elteres_fuvarok) > 20 else '')
        )

    df['Utolsó Leadási állomás időkapu (dátum)'] = pd.to_datetime(
        df['Utolsó Leadási állomás időkapu (dátum)'], errors='coerce')
    df['Első Felvételi állomás időkapu (dátum)'] = pd.to_datetime(
        df['Első Felvételi állomás időkapu (dátum)'], errors='coerce')

    # Dátum-anomália ellenőrzés: leadás < felvétel (adatrögzítési hiba, a
    # kör-építés rendezését keverheti)
    _anomalia_mask = (
        df['Utolsó Leadási állomás időkapu (dátum)'].notna()
        & df['Első Felvételi állomás időkapu (dátum)'].notna()
        & (df['Utolsó Leadási állomás időkapu (dátum)']
           < df['Első Felvételi állomás időkapu (dátum)'])
    )
    # v4.7: a cím-gazetteert már a riportok előtt fel kell építeni
    build_address_gazetteer(df)

    datum_anomalia_df = df.loc[_anomalia_mask, [
        'Fuvarszám', 'Járatszám', 'Vontatmány', 'Megbízó',
        'Első Felvételi állomás időkapu (dátum)',
        'Utolsó Leadási állomás időkapu (dátum)',
    ]].copy() if _anomalia_mask.any() else pd.DataFrame()
    if _anomalia_mask.any():
        st.warning(
            f'⚠️ {int(_anomalia_mask.sum())} sorban a leadási időkapu KORÁBBI, mint a '
            f'felvételi (adatrögzítési hiba gyanú). Ezek a kör-építés sorrendjét torzíthatják.'
        )
        with st.expander('Dátum-anomáliás sorok megtekintése'):
            st.dataframe(datum_anomalia_df, use_container_width=True)

    # v4.7 riportok: részfeladat-sorrend eltérés + feloldatlan címek
    reszfeladat_sorrend_df = build_reszfeladat_sorrend_table(df)
    if not reszfeladat_sorrend_df.empty:
        st.info(
            f'ℹ️ {len(reszfeladat_sorrend_df)} törzsnél a részfeladat-sorszámok sorrendje '
            f'eltér az időkapuk sorrendjétől. A generálás ezt kezeli (nem ad rá hamis '
            f'típushibát), de a forrásrendszerben javítandó — lásd a '
            f'"Részfeladat-sorrend eltérés" munkalapot.'
        )
    feloldatlan_cim_df = unresolved_addresses(df)
    if not feloldatlan_cim_df.empty:
        st.info(
            f'ℹ️ {len(feloldatlan_cim_df)} címnél az ország nem volt feloldható '
            f'(hiányzó országkód és ismeretlen település) — lásd a "Feloldatlan címek" '
            f'munkalapot.'
        )

    # Elérhető év-hónap párok a leadási dátumok alapján
    _kv_dates = df['Utolsó Leadási állomás időkapu (dátum)'].dropna()
    _available_ym = sorted({(int(d.year), int(d.month)) for d in _kv_dates})
    _ym_options = [f'{y}-{m:02d}' for y, m in _available_ym]

    selected_ym = st.multiselect(
        'Válaszd ki a legyártandó hónapokat (év-hónap) – mindegyikről külön Excel készül',
        _ym_options,
        default=_ym_options,
    )
    make_summary = st.checkbox(
        '📊 Összesítő Excel készítése a legyártott hónapokról', value=True)

    if st.button('🔄 Körfuvarok generálása', type='primary'):
        if not selected_ym:
            st.error('Válassz ki legalább egy hónapot!')
        else:
            with st.container():
                # --- Feldolgozási animáció + folyamatjelző ---
                _anim_ph = st.empty()
                _status_ph = st.empty()
                _progress = st.progress(0)
                _anim_ph.markdown(TRUCK_ANIM_HTML, unsafe_allow_html=True)

                def _status(txt, pct):
                    _status_ph.markdown(f'🛣️ **{txt}**')
                    _progress.progress(int(min(max(pct, 0), 100)))

                _status('Fuvarnapló elemzése, irányok osztályozása…', 5)
                # A cím-gazetteert az irány-osztályozás ELŐTT kell felépíteni (v4.7)
                build_address_gazetteer(df)
                df['Irány'] = df.apply(classify_leg_direction, axis=1)

                # Kör-építés EGYSZER, a teljes adathalmazon (a hónap-szűrés csak ezután
                # jön, így a hónaphatáron átnyúló körök helyesen épülnek fel)
                _status('Körfuvarok építése és összefűzése…', 15)
                result_df_all = generate_result_df(df)
                # a sorrend-riport Hatás oszlopa csak a generálás után tölthető ki
                reszfeladat_sorrend_df = build_reszfeladat_sorrend_table(df)
                kor_vege_ser = pd.to_datetime(result_df_all['Kör vége dátum'], errors='coerce')

                # Költség- és eredménykimutatás fájlok beolvasása EGYSZER
                _status('Költség- és eredménykimutatás fájlok beolvasása…', 35)
                combined_cost_df, cost_categories = load_all_cost_files(cost_files)
                has_cost = combined_cost_df is not None and not combined_cost_df.empty
                if not has_cost:
                    cost_categories = []
                combined_fc_df = load_all_flight_controlling_files(fc_files)
                has_fc = combined_fc_df is not None and not combined_fc_df.empty
                if has_cost:
                    st.info(
                        f'💶 Költség fájl(ok) feldolgozva: {len(combined_cost_df)} egyedi járat, '
                        f'{len(cost_categories)} költség kategória.'
                    )
                if has_fc:
                    st.info(
                        f'🛣️ Eredménykimutatás fájl(ok) feldolgozva: {len(combined_fc_df)} egyedi járat.'
                    )

                # Körfuvar-only mód: ha SEM költség, SEM eredménykimutatás fájl nincs
                # feltöltve, a feldolgozás csak a körfuvarok generálására terjed ki
                # (nincs költség/futási aggregálás, Teljes költség, Járati eredmény,
                # ország-reláció és megbízó profit-elemzés) → gyorsabb futás a
                # rögzítési hibák javítási köreihez.
                korfuvar_only = (not has_cost) and (not has_fc)
                if korfuvar_only:
                    st.info(
                        '⚡ Körfuvar-only mód: nincs feltöltve költség / eredménykimutatás '
                        'fájl, ezért csak a körfuvar-generálás és a hibalisták készülnek el '
                        '(költség- és futásadat-aggregálás kihagyva).'
                    )

                outputs = {}
                month_views = {}
                month_results = {}
                tipushiba_frames = []

                _n_sel = len(selected_ym)
                for _i_month, label in enumerate(selected_ym, 1):
                    _status(f'{label} körfuvar-tábla generálása ({_i_month}/{_n_sel})…',
                            40 + int(50 * (_i_month - 1) / _n_sel))
                    sel_year, sel_month = (int(x) for x in label.split('-'))
                    mask = (kor_vege_ser.dt.year == sel_year) & (kor_vege_ser.dt.month == sel_month)
                    result_df = result_df_all[mask].reset_index(drop=True)
                    if result_df.empty:
                        st.warning(f'⚠️ {label}: nincs kör ebben a hónapban, kihagyva.')
                        continue

                    if not korfuvar_only:
                        if has_cost:
                            result_df = aggregate_cost_for_rings(
                                result_df, combined_cost_df, cost_categories)
                        if has_fc:
                            result_df = aggregate_flight_controlling_for_rings(
                                result_df, combined_fc_df)

                        all_cost_cols = list(cost_categories) if has_cost else []
                        if has_fc:
                            for c in FC_EXTRA_COST_COLS:
                                if c in result_df.columns and c not in all_cost_cols:
                                    all_cost_cols.append(c)
                        result_df = finalize_totals(result_df, all_cost_cols)

                    # Elemző táblák ehhez a hónaphoz (körfuvar-only módban csak a
                    # hibajavításhoz szükségesek készülnek el)
                    tipushiba_df = build_tipushiba_table(result_df, df)
                    osszesito_df = build_osszesito_table(result_df)
                    if korfuvar_only:
                        relacio_df = pd.DataFrame()
                        megbizo_df = pd.DataFrame()
                    else:
                        relacio_df = build_relacio_table(result_df)
                        megbizo_df = build_megbizo_table(result_df)
                    if not tipushiba_df.empty:
                        _th = tipushiba_df.copy()
                        _th.insert(0, 'Hónap', label)
                        tipushiba_frames.append(_th)

                    # Oszlop-sorrend + csoport-színek
                    ordered_cols, group_of_col, futas_cols, koltseg_cols = \
                        order_and_group_columns(result_df, cost_categories)
                    display_cols = [c for c in ordered_cols if c in result_df.columns]
                    result_df_display = result_df[display_cols + ['Magyarázat_szín']].copy()

                    sum_cols = set(
                        ['Összes díj részarány (EUR)', 'Teljes költség', 'Járati eredmény', 'Km']
                        + futas_cols + koltseg_cols
                    )
                    xlsx_bytes = build_month_xlsx(
                        result_df_display, group_of_col, sum_cols,
                        extra_sheets=[
                            (osszesito_df, 'Összesítő'),
                            (relacio_df, 'Ország-relációk'),
                            (megbizo_df, 'Megbízók'),
                            (tipushiba_df, 'Típushibák (javítandó)'),
                            (datum_anomalia_df, 'Dátum-anomáliák'),
                            (reszfeladat_sorrend_df, 'Részfeladat-sorrend eltérés'),
                            (feloldatlan_cim_df, 'Feloldatlan címek'),
                        ],
                    )
                    outputs[f'korfuvarok_{sel_year}_{sel_month}.xlsx'] = xlsx_bytes
                    month_views[label] = {
                        'display': result_df_display,
                        'group_of_col': group_of_col,
                    }
                    month_results[label] = result_df

                # Összevont típushiba-lista (törzsenként egyszer)
                if tipushiba_frames:
                    tipushiba_all = pd.concat(tipushiba_frames, ignore_index=True)
                    tipushiba_all = tipushiba_all.drop_duplicates(
                        subset='Törzs', keep='first').reset_index(drop=True)
                else:
                    tipushiba_all = pd.DataFrame()

                # Opcionális összesítő munkafüzet a legyártott hónapokról
                _status('Összesítő és letölthető fájlok készítése…', 92)
                if make_summary and month_results:
                    _labels_sorted = sorted(month_results.keys())
                    _sum_name = (
                        f'korfuvarok_osszesito_{_labels_sorted[0]}_{_labels_sorted[-1]}.xlsx'
                        if len(_labels_sorted) > 1
                        else f'korfuvarok_osszesito_{_labels_sorted[0]}.xlsx'
                    )
                    outputs[_sum_name] = build_summary_xlsx(
                        month_results, tipushiba_all, datum_anomalia_df)

                # Session state-be mentjük, hogy a letöltés-gombok kattintása után
                # (Streamlit rerun) is megmaradjanak az eredmények
                st.session_state['kf_outputs'] = outputs
                st.session_state['kf_month_views'] = month_views
                st.session_state['kf_havi_osszesito'] = (
                    build_havi_osszesito_table(month_results) if month_results else pd.DataFrame()
                )
                st.session_state['kf_tipushiba_all'] = tipushiba_all

                # Animáció leszedése + siker jelzés
                _status('Kész!', 100)
                _anim_ph.empty()
                _status_ph.empty()
                _progress.empty()
                if outputs:
                    st.balloons()
                    st.success(f'✅ Feldolgozás kész: {len(outputs)} fájl legyártva.')

    # --- Eredmények megjelenítése (session_state-ből, letöltés után is megmarad) ---
    if st.session_state.get('kf_outputs'):
        st.markdown('---')
        st.subheader('📥 Legyártott fájlok')
        for _fname, _data in st.session_state['kf_outputs'].items():
            st.download_button(
                label=f'📥 {_fname}',
                data=_data,
                file_name=_fname,
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                key=f'dl_{_fname}',
            )

        _havi = st.session_state.get('kf_havi_osszesito')
        if _havi is not None and len(_havi):
            st.subheader('Havi összesítő')
            st.dataframe(_havi, use_container_width=True)

        _tipushiba_all = st.session_state.get('kf_tipushiba_all')
        if _tipushiba_all is not None and len(_tipushiba_all):
            with st.expander(
                    f'📋 Forrásrendszerben javítandó típus-/állomáshibák '
                    f'({len(_tipushiba_all)} törzs)'):
                st.dataframe(_tipushiba_all, use_container_width=True)

        _views = st.session_state.get('kf_month_views', {})
        for _label, _view in _views.items():
            _disp = _view['display']
            _group_of_col = _view['group_of_col']
            with st.expander(
                    f'Körfuvarok – {_label} ({len(_disp)} kör)',
                    expanded=(len(_views) == 1)):

                def _highlight_cells(row):
                    styles = []
                    for col in row.index:
                        if col == 'Magyarázat':
                            styles.append(row.get('Magyarázat_szín', '') or '')
                        elif col == 'Járati eredmény':
                            v = row.get('Járati eredmény')
                            if pd.notna(v):
                                if v < 0:
                                    styles.append('background-color: lightcoral; color: black')
                                elif v > 0:
                                    styles.append('background-color: lightgreen; color: black')
                                else:
                                    styles.append('')
                            else:
                                styles.append('')
                        else:
                            styles.append('')
                    return styles

                # v4.8 – teljesítmény: nagy táblánál nem rajzolunk Styler-t, mert a
                # soronkénti HTML-generálás több ezer sornál használhatatlanná teszi
                # az oldalt. A teljes tábla a letöltött Excelben van.
                _full_len = len(_disp)
                _show_all = False
                if _full_len > UI_PREVIEW_ROWS:
                    _show_all = st.checkbox(
                        f'Teljes tábla megjelenítése ({_full_len} sor) – lassú lehet',
                        value=False, key=f'showall_{_label}')
                _disp_ui = _disp if _show_all else _disp.head(UI_PREVIEW_ROWS)
                if not _show_all and _full_len > UI_PREVIEW_ROWS:
                    st.caption(
                        f'Előnézet: az első {UI_PREVIEW_ROWS} kör a(z) {_full_len}-ből. '
                        f'A teljes tábla a fenti letöltött Excelben van.')

                _shown = _disp_ui.drop(columns=['Magyarázat_szín'], errors='ignore')
                if len(_disp_ui) > UI_PREVIEW_ROWS:
                    # Színezés nélküli, virtualizált tábla – ez bírja a nagy adatot
                    st.dataframe(_shown, use_container_width=True)
                    continue
                try:
                    _styler = _disp_ui.style.apply(_highlight_cells, axis=1, subset=None)
                    if 'Magyarázat_szín' in _disp_ui.columns:
                        _styler = _styler.hide(axis='columns', subset=['Magyarázat_szín'])
                    _header_styles = []
                    for _i, _col in enumerate(list(_shown.columns)):
                        _g = _group_of_col.get(_col)
                        if _g:
                            _header_styles.append({
                                'selector': f'th.col_heading.level0.col{_i}',
                                'props': GROUP_COLORS[_g]['css'] + '; font-weight: bold;',
                            })
                    if _header_styles:
                        _styler = _styler.set_table_styles(_header_styles, overwrite=False)
                    st.dataframe(_styler, use_container_width=True)
                except Exception:
                    # Fallback: színezés nélküli tábla, ha a Styler nem elérhető
                    st.dataframe(_shown, use_container_width=True)
